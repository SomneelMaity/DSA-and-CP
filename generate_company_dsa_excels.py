"""
generate_company_dsa_excels.py
──────────────────────────────
Generates individual DSA prep Excel workbooks for:
  Intuit, Adobe, Microsoft, Amazon, JP Morgan, Morgan Stanley,
  Nvidia, Rubrik, Databricks, Jane Street, LinkedIn, Snowflake,
  Meta, Apple, HRT, OpenAI

Each workbook has:
  • Overview sheet  — company tips + sheet directory
  • Topic sheets    — one per DSA topic (Arrays, Trees, DP, …)
  • High Frequency  — company-specific top asked questions
"""

import os
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Output directory ─────────────────────────────────────────────────────────
OUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Company_DSA_Excels")
os.makedirs(OUT_DIR, exist_ok=True)

# ─── Shared Styles ────────────────────────────────────────────────────────────
DIFF_FILL = {
    "Easy":   PatternFill("solid", fgColor="C8E6C9"),
    "Medium": PatternFill("solid", fgColor="FFF9C4"),
    "Hard":   PatternFill("solid", fgColor="FFCDD2"),
}
WHITE  = Font(color="FFFFFF", bold=True, size=11)
BOLD   = Font(bold=True, size=10)
NORMAL = Font(size=10)
LINK_COLOR = "1565C0"
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
THIN   = Side(style="thin", color="BDBDBD")
BOX    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ─── Helpers ─────────────────────────────────────────────────────────────────
def safe_title(name):
    for ch in r"/\?*[]:'":
        name = name.replace(ch, "-")
    return name[:31]


def col_hdr_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def title_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def add_topic_sheet(wb, topic_name, rows, company_color, company_name):
    """
    rows: list of (no, problem, difficulty, lc_number, url, notes)
    Columns: #, Problem, Difficulty, LC#, LeetCode Link, Notes/Pattern, Status
    """
    ws = wb.create_sheet(title=safe_title(topic_name))

    # Title
    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value     = f"{company_name} DSA Prep — {topic_name}"
    c.fill      = title_fill(company_color)
    c.font      = Font(color="FFFFFF", bold=True, size=13)
    c.alignment = CENTER
    ws.row_dimensions[1].height = 28

    # Column headers
    headers    = ["#", "Problem", "Difficulty", "LC #", "LeetCode Link", "Notes / Pattern", "Status"]
    col_widths = [5,   52,        12,           8,      46,              36,                 14]
    hdr_fill   = col_hdr_fill(company_color)
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.fill      = hdr_fill
        cell.font      = WHITE
        cell.alignment = CENTER
        cell.border    = BOX
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[2].height = 20

    # Data rows
    for r_idx, (no, problem, difficulty, lc_num, url, notes) in enumerate(rows, start=3):
        fill = DIFF_FILL.get(difficulty, PatternFill())
        ws.cell(row=r_idx, column=1, value=no).alignment         = CENTER
        ws.cell(row=r_idx, column=2, value=problem).alignment    = LEFT
        ws.cell(row=r_idx, column=3, value=difficulty).alignment = CENTER
        ws.cell(row=r_idx, column=4, value=lc_num).alignment     = CENTER

        link_cell = ws.cell(row=r_idx, column=5, value=f"LC {lc_num} — Open")
        link_cell.hyperlink  = url
        link_cell.font       = Font(color=LINK_COLOR, underline="single", size=10)
        link_cell.alignment  = LEFT

        ws.cell(row=r_idx, column=6, value=notes).alignment = LEFT
        ws.cell(row=r_idx, column=6).font = Font(size=9, color="555555", italic=True)

        status_cell = ws.cell(row=r_idx, column=7, value="⬜ To Do")
        status_cell.alignment = CENTER
        status_cell.font      = Font(size=9, color="555555")

        for col in range(1, 8):
            cell = ws.cell(row=r_idx, column=col)
            cell.fill   = fill
            cell.border = BOX
            if col not in (5, 6, 7):
                cell.font = NORMAL
        ws.row_dimensions[r_idx].height = 18

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:G{2 + len(rows)}"
    return ws


def add_hf_sheet(wb, company_name, company_color, hf_rows):
    """
    hf_rows: list of (no, problem, difficulty, lc_number, url, why_asked)
    High-Frequency company-specific questions sheet.
    """
    ws = wb.create_sheet(title="⭐ High Frequency")

    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value     = f"{company_name} — Top High-Frequency Questions"
    c.fill      = PatternFill("solid", fgColor="B71C1C")
    c.font      = Font(color="FFFFFF", bold=True, size=13)
    c.alignment = CENTER
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:G2")
    d = ws["A2"]
    d.value     = "These questions appear most frequently in actual interviews. Prioritise solving them first!"
    d.fill      = PatternFill("solid", fgColor="FFF8E1")
    d.font      = Font(italic=True, size=10, color="333333")
    d.alignment = LEFT
    ws.row_dimensions[2].height = 18

    headers    = ["#", "Problem", "Difficulty", "LC #", "LeetCode Link", "Why Asked / Pattern", "Status"]
    col_widths = [5,   52,        12,           8,      46,              38,                     14]
    hdr_fill   = col_hdr_fill(company_color)
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.fill      = hdr_fill
        cell.font      = WHITE
        cell.alignment = CENTER
        cell.border    = BOX
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[3].height = 20

    for r_idx, (no, problem, difficulty, lc_num, url, why) in enumerate(hf_rows, start=4):
        fill = DIFF_FILL.get(difficulty, PatternFill())
        ws.cell(row=r_idx, column=1, value=no).alignment         = CENTER
        ws.cell(row=r_idx, column=2, value=problem).alignment    = LEFT
        ws.cell(row=r_idx, column=3, value=difficulty).alignment = CENTER
        ws.cell(row=r_idx, column=4, value=lc_num).alignment     = CENTER

        link_cell = ws.cell(row=r_idx, column=5, value=f"LC {lc_num} — Open")
        link_cell.hyperlink  = url
        link_cell.font       = Font(color=LINK_COLOR, underline="single", size=10)
        link_cell.alignment  = LEFT

        ws.cell(row=r_idx, column=6, value=why).alignment = LEFT
        ws.cell(row=r_idx, column=6).font = Font(size=9, color="333333", italic=True, bold=True)

        status_cell = ws.cell(row=r_idx, column=7, value="⬜ To Do")
        status_cell.alignment = CENTER
        status_cell.font      = Font(size=9, color="555555")

        for col in range(1, 8):
            cell = ws.cell(row=r_idx, column=col)
            cell.fill   = fill
            cell.border = BOX
            if col not in (5, 6, 7):
                cell.font = NORMAL
        ws.row_dimensions[r_idx].height = 18

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:G{3 + len(hf_rows)}"
    return ws


def add_overview_sheet(wb, company_name, company_color, tips, topic_names_and_counts):
    ws = wb.create_sheet(title="Overview", index=0)

    # Banner
    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value     = f"{company_name} — DSA Interview Preparation"
    c.fill      = title_fill(company_color)
    c.font      = Font(color="FFFFFF", bold=True, size=15)
    c.alignment = CENTER
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:E2")
    sub = ws["A2"]
    sub.value     = "Topic-wise question bank with status tracker | ⭐ High Frequency sheet is your PRIORITY"
    sub.fill      = PatternFill("solid", fgColor="FFF8E1")
    sub.font      = Font(italic=True, size=10, color="333333")
    sub.alignment = CENTER
    ws.row_dimensions[2].height = 20

    # Tips section
    ws.merge_cells("A4:E4")
    tip_hdr = ws["A4"]
    tip_hdr.value     = f"  {company_name} Interview Tips"
    tip_hdr.fill      = col_hdr_fill(company_color)
    tip_hdr.font      = WHITE
    tip_hdr.alignment = LEFT
    tip_hdr.border    = BOX
    ws.row_dimensions[4].height = 22

    for i, tip in enumerate(tips, start=5):
        ws.merge_cells(f"A{i}:E{i}")
        cell = ws.cell(row=i, column=1, value=f"  • {tip}")
        cell.fill      = PatternFill("solid", fgColor="FFF8E1")
        cell.font      = Font(size=10, italic=True, color="333333")
        cell.alignment = LEFT
        cell.border    = BOX
        ws.row_dimensions[i].height = 22

    # Directory table
    tbl_start = 5 + len(tips) + 1
    ws.merge_cells(f"A{tbl_start}:E{tbl_start}")
    tbl_hdr = ws.cell(row=tbl_start, column=1, value="  Sheet Directory")
    tbl_hdr.fill      = col_hdr_fill(company_color)
    tbl_hdr.font      = WHITE
    tbl_hdr.alignment = LEFT
    tbl_hdr.border    = BOX
    ws.row_dimensions[tbl_start].height = 22

    sub_hdrs   = ["Sheet Name", "# Questions", "Topics Covered", "Priority", "Notes"]
    col_widths = [36,           14,             32,               14,          28]
    for col, (h, w) in enumerate(zip(sub_hdrs, col_widths), 1):
        cell = ws.cell(row=tbl_start + 1, column=col, value=h)
        cell.fill      = PatternFill("solid", fgColor="42A5F5")
        cell.font      = WHITE
        cell.alignment = CENTER
        cell.border    = BOX
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[tbl_start + 1].height = 20

    ALT1 = PatternFill("solid", fgColor="E3F2FD")
    ALT2 = PatternFill("solid", fgColor="FFFFFF")
    HF   = PatternFill("solid", fgColor="FFCDD2")

    for r_idx, (sheet_name, q_count, topics_str, priority, note) in enumerate(topic_names_and_counts, start=tbl_start + 2):
        fill = HF if "High Frequency" in sheet_name else (ALT1 if r_idx % 2 == 0 else ALT2)
        vals = [sheet_name, q_count, topics_str, priority, note]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=r_idx, column=col, value=val)
            cell.fill      = fill
            cell.font      = BOLD if col == 1 else NORMAL
            cell.alignment = LEFT if col != 2 else CENTER
            cell.border    = BOX
        ws.row_dimensions[r_idx].height = 18

    ws.freeze_panes = f"A{tbl_start + 2}"
    return ws


# ═══════════════════════════════════════════════════════════════════════════════
#  SHARED TOPIC QUESTION BANK
#  Format: (no, problem, difficulty, lc_number, url, notes)
# ═══════════════════════════════════════════════════════════════════════════════

TOPICS = {}

TOPICS["Arrays & Hashing"] = [
    (1,  "Two Sum",                               "Easy",   1,    "https://leetcode.com/problems/two-sum/",                               "Hash map; complement lookup"),
    (2,  "Contains Duplicate",                    "Easy",   217,  "https://leetcode.com/problems/contains-duplicate/",                   "Hash set"),
    (3,  "Valid Anagram",                         "Easy",   242,  "https://leetcode.com/problems/valid-anagram/",                        "Char frequency"),
    (4,  "Group Anagrams",                        "Medium", 49,   "https://leetcode.com/problems/group-anagrams/",                       "Sorted key or char-count tuple"),
    (5,  "Top K Frequent Elements",               "Medium", 347,  "https://leetcode.com/problems/top-k-frequent-elements/",              "Heap or bucket sort"),
    (6,  "Product of Array Except Self",          "Medium", 238,  "https://leetcode.com/problems/product-of-array-except-self/",         "Prefix & suffix products; no division"),
    (7,  "Valid Sudoku",                          "Medium", 36,   "https://leetcode.com/problems/valid-sudoku/",                         "Hash sets per row/col/box"),
    (8,  "Longest Consecutive Sequence",          "Medium", 128,  "https://leetcode.com/problems/longest-consecutive-sequence/",         "Hash set; O(n)"),
    (9,  "Subarray Sum Equals K",                 "Medium", 560,  "https://leetcode.com/problems/subarray-sum-equals-k/",                "Prefix sum + hash map"),
    (10, "Sort Colors (Dutch Flag)",              "Medium", 75,   "https://leetcode.com/problems/sort-colors/",                          "3-pointer in-place"),
    (11, "Find All Duplicates in Array",          "Medium", 442,  "https://leetcode.com/problems/find-all-duplicates-in-an-array/",      "Negate-at-index trick"),
    (12, "Maximum Points from Cards",             "Medium", 1423, "https://leetcode.com/problems/maximum-points-you-can-obtain-from-cards/", "Sliding window fixed"),
    (13, "First Missing Positive",                "Hard",   41,   "https://leetcode.com/problems/first-missing-positive/",               "Index-as-hash; cyclic sort"),
    (14, "Trapping Rain Water",                   "Hard",   42,   "https://leetcode.com/problems/trapping-rain-water/",                   "Two pointers or monotonic stack"),
    (15, "Largest Rectangle in Histogram",        "Hard",   84,   "https://leetcode.com/problems/largest-rectangle-in-histogram/",       "Monotonic stack"),
]

TOPICS["Two Pointers"] = [
    (1,  "Valid Palindrome",                     "Easy",   125,  "https://leetcode.com/problems/valid-palindrome/",                     "Skip non-alnum"),
    (2,  "Two Sum II — Sorted Array",            "Medium", 167,  "https://leetcode.com/problems/two-sum-ii-input-array-is-sorted/",    "Left/right converge"),
    (3,  "3Sum",                                 "Medium", 15,   "https://leetcode.com/problems/3sum/",                                 "Sort + two pointers"),
    (4,  "4Sum",                                 "Medium", 18,   "https://leetcode.com/problems/4sum/",                                 "Reduce to 3Sum"),
    (5,  "Container With Most Water",            "Medium", 11,   "https://leetcode.com/problems/container-with-most-water/",            "Greedy shrink shorter side"),
    (6,  "Move Zeroes",                          "Easy",   283,  "https://leetcode.com/problems/move-zeroes/",                          "In-place; slow pointer"),
    (7,  "Remove Duplicates from Sorted Array",  "Easy",   26,   "https://leetcode.com/problems/remove-duplicates-from-sorted-array/",  "Write pointer"),
    (8,  "Trapping Rain Water",                  "Hard",   42,   "https://leetcode.com/problems/trapping-rain-water/",                   "Two pointers; track max L/R"),
    (9,  "Boats to Save People",                 "Medium", 881,  "https://leetcode.com/problems/boats-to-save-people/",                 "Greedy; sort + two ends"),
    (10, "Minimum Size Subarray Sum",            "Medium", 209,  "https://leetcode.com/problems/minimum-size-subarray-sum/",            "Variable window"),
]

TOPICS["Sliding Window"] = [
    (1,  "Best Time to Buy and Sell Stock",            "Easy",   121, "https://leetcode.com/problems/best-time-to-buy-and-sell-stock/",               "Track min so far"),
    (2,  "Longest Substring Without Repeating",        "Medium", 3,   "https://leetcode.com/problems/longest-substring-without-repeating-characters/","Sliding window + set"),
    (3,  "Longest Repeating Character Replacement",    "Medium", 424, "https://leetcode.com/problems/longest-repeating-character-replacement/",        "Max-freq trick"),
    (4,  "Permutation in String",                      "Medium", 567, "https://leetcode.com/problems/permutation-in-string/",                          "Fixed window + char count"),
    (5,  "Find All Anagrams in a String",              "Medium", 438, "https://leetcode.com/problems/find-all-anagrams-in-a-string/",                  "Fixed sliding window"),
    (6,  "Minimum Window Substring",                   "Hard",   76,  "https://leetcode.com/problems/minimum-window-substring/",                       "Shrink when valid"),
    (7,  "Sliding Window Maximum",                     "Hard",   239, "https://leetcode.com/problems/sliding-window-maximum/",                          "Monotonic deque"),
    (8,  "Fruit Into Baskets",                         "Medium", 904, "https://leetcode.com/problems/fruit-into-baskets/",                              "At most 2 distinct"),
    (9,  "Longest Substring K Distinct Chars",         "Medium", 340, "https://leetcode.com/problems/longest-substring-with-at-most-k-distinct-characters/", "Map + shrink"),
    (10, "Max Consecutive Ones III",                   "Medium", 1004,"https://leetcode.com/problems/max-consecutive-ones-iii/",                        "At most K zeros"),
]

TOPICS["Stack"] = [
    (1,  "Valid Parentheses",                    "Easy",   20,   "https://leetcode.com/problems/valid-parentheses/",                          "Classic stack"),
    (2,  "Min Stack",                            "Medium", 155,  "https://leetcode.com/problems/min-stack/",                                  "Pair-stack or aux stack"),
    (3,  "Evaluate Reverse Polish Notation",     "Medium", 150,  "https://leetcode.com/problems/evaluate-reverse-polish-notation/",           "Operand stack"),
    (4,  "Generate Parentheses",                 "Medium", 22,   "https://leetcode.com/problems/generate-parentheses/",                       "Backtracking with open/close counts"),
    (5,  "Daily Temperatures",                   "Medium", 739,  "https://leetcode.com/problems/daily-temperatures/",                         "Monotonic decreasing stack"),
    (6,  "Car Fleet",                            "Medium", 853,  "https://leetcode.com/problems/car-fleet/",                                  "Simulate with stack"),
    (7,  "Largest Rectangle in Histogram",       "Hard",   84,   "https://leetcode.com/problems/largest-rectangle-in-histogram/",             "Monotonic increasing stack"),
    (8,  "Asteroid Collision",                   "Medium", 735,  "https://leetcode.com/problems/asteroid-collision/",                         "Stack simulation"),
    (9,  "Minimum Remove to Make Valid Parens",  "Medium", 1249, "https://leetcode.com/problems/minimum-remove-to-make-valid-parentheses/",   "Stack of indices"),
    (10, "Decode String",                        "Medium", 394,  "https://leetcode.com/problems/decode-string/",                              "Stack for count/string pairs"),
    (11, "Remove Duplicate Letters",             "Medium", 316,  "https://leetcode.com/problems/remove-duplicate-letters/",                   "Greedy + monotonic stack"),
    (12, "Basic Calculator II",                  "Medium", 227,  "https://leetcode.com/problems/basic-calculator-ii/",                        "Stack; handle * /"),
    (13, "Basic Calculator",                     "Hard",   224,  "https://leetcode.com/problems/basic-calculator/",                           "Stack for sign context"),
]

TOPICS["Binary Search"] = [
    (1,  "Binary Search",                             "Easy",   704,  "https://leetcode.com/problems/binary-search/",                                           "Template"),
    (2,  "Search a 2D Matrix",                        "Medium", 74,   "https://leetcode.com/problems/search-a-2d-matrix/",                                      "Treat as 1D sorted"),
    (3,  "Koko Eating Bananas",                       "Medium", 875,  "https://leetcode.com/problems/koko-eating-bananas/",                                     "Binary search on answer"),
    (4,  "Find Minimum in Rotated Sorted Array",      "Medium", 153,  "https://leetcode.com/problems/find-minimum-in-rotated-sorted-array/",                   "Binary search"),
    (5,  "Search in Rotated Sorted Array",            "Medium", 33,   "https://leetcode.com/problems/search-in-rotated-sorted-array/",                         "Modified binary search"),
    (6,  "Time Based Key-Value Store",                "Medium", 981,  "https://leetcode.com/problems/time-based-key-value-store/",                              "Binary search on timestamps"),
    (7,  "Find First and Last Position",              "Medium", 34,   "https://leetcode.com/problems/find-first-and-last-position-of-element-in-sorted-array/","Two binary searches"),
    (8,  "Median of Two Sorted Arrays",               "Hard",   4,    "https://leetcode.com/problems/median-of-two-sorted-arrays/",                             "Binary search on partition"),
    (9,  "Capacity to Ship in D Days",                "Medium", 1011, "https://leetcode.com/problems/capacity-to-ship-packages-within-d-days/",                "Binary search on answer"),
    (10, "Split Array Largest Sum",                   "Hard",   410,  "https://leetcode.com/problems/split-array-largest-sum/",                                 "Binary search on answer"),
    (11, "Peak Index in Mountain Array",              "Medium", 852,  "https://leetcode.com/problems/peak-index-in-a-mountain-array/",                         "Binary search on peak"),
    (12, "Path With Minimum Effort",                  "Medium", 1631, "https://leetcode.com/problems/path-with-minimum-effort/",                                "Binary search + BFS"),
]

TOPICS["Linked List"] = [
    (1,  "Reverse Linked List",                "Easy",   206, "https://leetcode.com/problems/reverse-linked-list/",                        "Iterative & recursive"),
    (2,  "Merge Two Sorted Lists",             "Easy",   21,  "https://leetcode.com/problems/merge-two-sorted-lists/",                     "Dummy head"),
    (3,  "Linked List Cycle",                  "Easy",   141, "https://leetcode.com/problems/linked-list-cycle/",                          "Floyd's slow/fast"),
    (4,  "Linked List Cycle II",               "Medium", 142, "https://leetcode.com/problems/linked-list-cycle-ii/",                       "Floyd's + math"),
    (5,  "Remove Nth Node From End",           "Medium", 19,  "https://leetcode.com/problems/remove-nth-node-from-end-of-list/",          "Two pointers gap n"),
    (6,  "Copy List with Random Pointer",      "Medium", 138, "https://leetcode.com/problems/copy-list-with-random-pointer/",             "Hash map or weave"),
    (7,  "Add Two Numbers",                    "Medium", 2,   "https://leetcode.com/problems/add-two-numbers/",                            "Carry propagation"),
    (8,  "Find the Duplicate Number",          "Medium", 287, "https://leetcode.com/problems/find-the-duplicate-number/",                 "Floyd's cycle (index as pointer)"),
    (9,  "LRU Cache",                          "Medium", 146, "https://leetcode.com/problems/lru-cache/",                                  "Doubly-LL + hash map"),
    (10, "Merge K Sorted Lists",               "Hard",   23,  "https://leetcode.com/problems/merge-k-sorted-lists/",                      "Min-heap or divide & conquer"),
    (11, "Reverse Nodes in K-Group",           "Hard",   25,  "https://leetcode.com/problems/reverse-nodes-in-k-group/",                  "Reverse in chunks"),
    (12, "Reorder List",                       "Medium", 143, "https://leetcode.com/problems/reorder-list/",                               "Find mid + reverse + merge"),
    (13, "Sort List",                          "Medium", 148, "https://leetcode.com/problems/sort-list/",                                  "Merge sort on LL"),
    (14, "Palindrome Linked List",             "Easy",   234, "https://leetcode.com/problems/palindrome-linked-list/",                    "Find mid + reverse second half"),
    (15, "Intersection of Two Linked Lists",   "Easy",   160, "https://leetcode.com/problems/intersection-of-two-linked-lists/",          "Two pointers; swap heads"),
]

TOPICS["Trees"] = [
    (1,  "Invert Binary Tree",                         "Easy",   226,  "https://leetcode.com/problems/invert-binary-tree/",                                               "Recursive swap"),
    (2,  "Maximum Depth of Binary Tree",               "Easy",   104,  "https://leetcode.com/problems/maximum-depth-of-binary-tree/",                                    "DFS"),
    (3,  "Diameter of Binary Tree",                    "Easy",   543,  "https://leetcode.com/problems/diameter-of-binary-tree/",                                         "Height DFS; max left+right"),
    (4,  "Balanced Binary Tree",                       "Easy",   110,  "https://leetcode.com/problems/balanced-binary-tree/",                                            "Height check DFS"),
    (5,  "Same Tree",                                  "Easy",   100,  "https://leetcode.com/problems/same-tree/",                                                        "Parallel DFS"),
    (6,  "Subtree of Another Tree",                    "Easy",   572,  "https://leetcode.com/problems/subtree-of-another-tree/",                                         "isSameTree at each node"),
    (7,  "LCA of Binary Search Tree",                  "Medium", 235,  "https://leetcode.com/problems/lowest-common-ancestor-of-a-binary-search-tree/",                 "Use BST property"),
    (8,  "LCA of Binary Tree",                         "Medium", 236,  "https://leetcode.com/problems/lowest-common-ancestor-of-a-binary-tree/",                        "Post-order DFS"),
    (9,  "Binary Tree Level Order Traversal",          "Medium", 102,  "https://leetcode.com/problems/binary-tree-level-order-traversal/",                               "BFS with queue"),
    (10, "Binary Tree Right Side View",                "Medium", 199,  "https://leetcode.com/problems/binary-tree-right-side-view/",                                     "BFS last per level"),
    (11, "Count Good Nodes in Binary Tree",            "Medium", 1448, "https://leetcode.com/problems/count-good-nodes-in-binary-tree/",                                 "DFS; track running max"),
    (12, "Validate Binary Search Tree",                "Medium", 98,   "https://leetcode.com/problems/validate-binary-search-tree/",                                     "Min/max range DFS"),
    (13, "Kth Smallest Element in BST",                "Medium", 230,  "https://leetcode.com/problems/kth-smallest-element-in-a-bst/",                                  "In-order traversal"),
    (14, "Construct BT from Preorder & Inorder",       "Medium", 105,  "https://leetcode.com/problems/construct-binary-tree-from-preorder-and-inorder-traversal/",      "Divide at root"),
    (15, "Binary Tree Maximum Path Sum",               "Hard",   124,  "https://leetcode.com/problems/binary-tree-maximum-path-sum/",                                    "Post-order; global max"),
    (16, "Serialize and Deserialize Binary Tree",      "Hard",   297,  "https://leetcode.com/problems/serialize-and-deserialize-binary-tree/",                           "BFS or preorder"),
    (17, "Vertical Order Traversal of Binary Tree",    "Hard",   987,  "https://leetcode.com/problems/vertical-order-traversal-of-a-binary-tree/",                      "DFS with (col,row) keys"),
    (18, "All Nodes Distance K in Binary Tree",        "Medium", 863,  "https://leetcode.com/problems/all-nodes-distance-k-in-binary-tree/",                            "Graph-ify tree; BFS"),
    (19, "Path Sum II",                                "Medium", 113,  "https://leetcode.com/problems/path-sum-ii/",                                                      "DFS backtrack"),
    (20, "Flatten Binary Tree to Linked List",         "Medium", 114,  "https://leetcode.com/problems/flatten-binary-tree-to-linked-list/",                              "Morris right-thread"),
]

TOPICS["Heap / Priority Queue"] = [
    (1,  "Kth Largest Element in Stream",    "Easy",   703,  "https://leetcode.com/problems/kth-largest-element-in-a-stream/",     "Min-heap size k"),
    (2,  "Last Stone Weight",                "Easy",   1046, "https://leetcode.com/problems/last-stone-weight/",                   "Max-heap"),
    (3,  "K Closest Points to Origin",       "Medium", 973,  "https://leetcode.com/problems/k-closest-points-to-origin/",          "Max-heap size k"),
    (4,  "Kth Largest Element in Array",     "Medium", 215,  "https://leetcode.com/problems/kth-largest-element-in-an-array/",     "Quick select or min-heap"),
    (5,  "Task Scheduler",                   "Medium", 621,  "https://leetcode.com/problems/task-scheduler/",                      "Greedy; max-heap + cooldown"),
    (6,  "Design Twitter",                   "Medium", 355,  "https://leetcode.com/problems/design-twitter/",                      "Min-heap merge k lists"),
    (7,  "Find Median from Data Stream",     "Hard",   295,  "https://leetcode.com/problems/find-median-from-data-stream/",        "Two heaps (max + min)"),
    (8,  "IPO (Maximize Capital)",           "Hard",   502,  "https://leetcode.com/problems/ipo/",                                 "Two heaps; greedy"),
    (9,  "Top K Frequent Words",             "Medium", 692,  "https://leetcode.com/problems/top-k-frequent-words/",                "Min-heap with custom compare"),
    (10, "Ugly Number II",                   "Medium", 264,  "https://leetcode.com/problems/ugly-number-ii/",                     "Min-heap or three pointers"),
    (11, "Merge K Sorted Lists",             "Hard",   23,   "https://leetcode.com/problems/merge-k-sorted-lists/",               "Min-heap"),
    (12, "Maximum Profit Job Scheduling",    "Hard",   1235, "https://leetcode.com/problems/maximum-profit-in-job-scheduling/",   "Sort + DP + heap"),
]

TOPICS["Backtracking"] = [
    (1,  "Subsets",                              "Medium", 78,  "https://leetcode.com/problems/subsets/",                               "Include/exclude"),
    (2,  "Combination Sum",                      "Medium", 39,  "https://leetcode.com/problems/combination-sum/",                      "Reuse elements"),
    (3,  "Permutations",                         "Medium", 46,  "https://leetcode.com/problems/permutations/",                         "Swap or used[] array"),
    (4,  "Subsets II",                           "Medium", 90,  "https://leetcode.com/problems/subsets-ii/",                           "Skip duplicates"),
    (5,  "Combination Sum II",                   "Medium", 40,  "https://leetcode.com/problems/combination-sum-ii/",                   "No reuse; skip dups"),
    (6,  "Word Search",                          "Medium", 79,  "https://leetcode.com/problems/word-search/",                          "DFS + visited mask"),
    (7,  "Palindrome Partitioning",              "Medium", 131, "https://leetcode.com/problems/palindrome-partitioning/",              "DFS + palindrome check"),
    (8,  "Letter Combinations of Phone Number",  "Medium", 17,  "https://leetcode.com/problems/letter-combinations-of-a-phone-number/","Map digit→chars; DFS"),
    (9,  "N-Queens",                             "Hard",   51,  "https://leetcode.com/problems/n-queens/",                             "Row by row; col/diag sets"),
    (10, "Sudoku Solver",                        "Hard",   37,  "https://leetcode.com/problems/sudoku-solver/",                        "Backtrack empty cells"),
    (11, "Expression Add Operators",             "Hard",   282, "https://leetcode.com/problems/expression-add-operators/",             "Track prev for * precedence"),
    (12, "Word Break II",                        "Hard",   140, "https://leetcode.com/problems/word-break-ii/",                        "Backtracking + memoisation"),
]

TOPICS["Tries"] = [
    (1,  "Implement Trie (Prefix Tree)",          "Medium", 208, "https://leetcode.com/problems/implement-trie-prefix-tree/",                  "TrieNode with children dict"),
    (2,  "Design Add and Search Words",           "Medium", 211, "https://leetcode.com/problems/design-add-and-search-words-data-structure/",  "DFS for '.' wildcard"),
    (3,  "Word Search II",                        "Hard",   212, "https://leetcode.com/problems/word-search-ii/",                              "Trie + DFS on board"),
    (4,  "Replace Words",                         "Medium", 648, "https://leetcode.com/problems/replace-words/",                               "Trie lookup"),
    (5,  "Maximum XOR of Two Numbers",            "Medium", 421, "https://leetcode.com/problems/maximum-xor-of-two-numbers-in-an-array/",      "Bit trie"),
    (6,  "Longest Word in Dictionary",            "Medium", 720, "https://leetcode.com/problems/longest-word-in-dictionary/",                  "Trie BFS"),
    (7,  "Map Sum Pairs",                         "Medium", 677, "https://leetcode.com/problems/map-sum-pairs/",                               "Trie with values"),
]

TOPICS["Graphs"] = [
    (1,  "Number of Islands",                     "Medium", 200,  "https://leetcode.com/problems/number-of-islands/",                       "BFS/DFS flood-fill"),
    (2,  "Clone Graph",                           "Medium", 133,  "https://leetcode.com/problems/clone-graph/",                             "BFS/DFS + hash map"),
    (3,  "Max Area of Island",                    "Medium", 695,  "https://leetcode.com/problems/max-area-of-island/",                      "DFS; count cells"),
    (4,  "Pacific Atlantic Water Flow",           "Medium", 417,  "https://leetcode.com/problems/pacific-atlantic-water-flow/",             "Reverse DFS from both oceans"),
    (5,  "Course Schedule",                       "Medium", 207,  "https://leetcode.com/problems/course-schedule/",                         "Topo sort / cycle detect"),
    (6,  "Course Schedule II",                    "Medium", 210,  "https://leetcode.com/problems/course-schedule-ii/",                      "Kahn's BFS topo sort"),
    (7,  "Accounts Merge",                        "Medium", 721,  "https://leetcode.com/problems/accounts-merge/",                          "Union-Find or DFS"),
    (8,  "Rotting Oranges",                       "Medium", 994,  "https://leetcode.com/problems/rotting-oranges/",                         "Multi-source BFS"),
    (9,  "Is Graph Bipartite?",                   "Medium", 785,  "https://leetcode.com/problems/is-graph-bipartite/",                      "2-colour BFS/DFS"),
    (10, "Word Ladder",                           "Hard",   127,  "https://leetcode.com/problems/word-ladder/",                             "BFS with word transforms"),
    (11, "Find Eventual Safe States",             "Medium", 802,  "https://leetcode.com/problems/find-eventual-safe-states/",               "Reverse graph topo sort"),
    (12, "Redundant Connection",                  "Medium", 684,  "https://leetcode.com/problems/redundant-connection/",                     "Union-Find; detect cycle"),
    (13, "Network Delay Time",                    "Medium", 743,  "https://leetcode.com/problems/network-delay-time/",                      "Dijkstra SSSP"),
    (14, "Cheapest Flights Within K Stops",       "Medium", 787,  "https://leetcode.com/problems/cheapest-flights-within-k-stops/",         "Bellman-Ford k+1 rounds"),
    (15, "All Paths From Source to Target",       "Medium", 797,  "https://leetcode.com/problems/all-paths-from-source-to-target/",         "DFS backtrack"),
    (16, "Critical Connections in Network",       "Hard",   1192, "https://leetcode.com/problems/critical-connections-in-a-network/",       "Tarjan's bridge"),
    (17, "Shortest Path in Binary Matrix",        "Medium", 1091, "https://leetcode.com/problems/shortest-path-in-binary-matrix/",          "BFS; 8-directional"),
    (18, "Min Vertices to Reach All Nodes",       "Medium", 1557, "https://leetcode.com/problems/minimum-number-of-vertices-to-reach-all-nodes/", "Nodes with 0 in-degree"),
    (19, "Swim in Rising Water",                  "Hard",   778,  "https://leetcode.com/problems/swim-in-rising-water/",                    "Dijkstra or binary search + BFS"),
    (20, "Reconstruct Itinerary",                 "Hard",   332,  "https://leetcode.com/problems/reconstruct-itinerary/",                   "Hierholzer's Eulerian path"),
    (21, "Evaluate Division",                     "Medium", 399,  "https://leetcode.com/problems/evaluate-division/",                       "Weighted graph BFS/DFS"),
    (22, "Min Cost to Connect All Points",        "Medium", 1584, "https://leetcode.com/problems/min-cost-to-connect-all-points/",          "Prim's / Kruskal's MST"),
]

TOPICS["DP — 1D"] = [
    (1,  "Climbing Stairs",                    "Easy",   70,   "https://leetcode.com/problems/climbing-stairs/",                   "Fibonacci DP"),
    (2,  "Min Cost Climbing Stairs",           "Easy",   746,  "https://leetcode.com/problems/min-cost-climbing-stairs/",         "Pick min of last two"),
    (3,  "House Robber",                       "Medium", 198,  "https://leetcode.com/problems/house-robber/",                     "No adjacent"),
    (4,  "House Robber II",                    "Medium", 213,  "https://leetcode.com/problems/house-robber-ii/",                  "Circular; two passes"),
    (5,  "Longest Palindromic Substring",      "Medium", 5,    "https://leetcode.com/problems/longest-palindromic-substring/",    "Expand around center"),
    (6,  "Palindromic Substrings",             "Medium", 647,  "https://leetcode.com/problems/palindromic-substrings/",           "Expand around center; count"),
    (7,  "Decode Ways",                        "Medium", 91,   "https://leetcode.com/problems/decode-ways/",                      "1D DP; handle leading zeros"),
    (8,  "Coin Change",                        "Medium", 322,  "https://leetcode.com/problems/coin-change/",                      "Unbounded knapsack"),
    (9,  "Maximum Product Subarray",           "Medium", 152,  "https://leetcode.com/problems/maximum-product-subarray/",         "Track min and max"),
    (10, "Word Break",                         "Medium", 139,  "https://leetcode.com/problems/word-break/",                       "BFS or 1D DP"),
    (11, "Longest Increasing Subsequence",     "Medium", 300,  "https://leetcode.com/problems/longest-increasing-subsequence/",   "DP O(n²) or patience sort O(n log n)"),
    (12, "Partition Equal Subset Sum",         "Medium", 416,  "https://leetcode.com/problems/partition-equal-subset-sum/",       "0/1 knapsack"),
    (13, "Jump Game",                          "Medium", 55,   "https://leetcode.com/problems/jump-game/",                        "Greedy max reach"),
    (14, "Jump Game II",                       "Medium", 45,   "https://leetcode.com/problems/jump-game-ii/",                     "Greedy levels"),
    (15, "Maximum Subarray (Kadane's)",        "Medium", 53,   "https://leetcode.com/problems/maximum-subarray/",                 "Kadane's algorithm"),
    (16, "Perfect Squares",                    "Medium", 279,  "https://leetcode.com/problems/perfect-squares/",                  "BFS or DP"),
    (17, "Fibonacci Number",                   "Easy",   509,  "https://leetcode.com/problems/fibonacci-number/",                 "Bottom-up or memo"),
]

TOPICS["DP — 2D"] = [
    (1,  "Unique Paths",                               "Medium", 62,   "https://leetcode.com/problems/unique-paths/",                                        "Grid DP"),
    (2,  "Minimum Path Sum",                           "Medium", 64,   "https://leetcode.com/problems/minimum-path-sum/",                                    "Grid DP; accumulate"),
    (3,  "Longest Common Subsequence",                 "Medium", 1143, "https://leetcode.com/problems/longest-common-subsequence/",                          "Classic 2D DP"),
    (4,  "Edit Distance",                              "Medium", 72,   "https://leetcode.com/problems/edit-distance/",                                        "Insert/delete/replace"),
    (5,  "Interleaving String",                        "Medium", 97,   "https://leetcode.com/problems/interleaving-string/",                                  "2D DP on two strings"),
    (6,  "Coin Change II",                             "Medium", 518,  "https://leetcode.com/problems/coin-change-ii/",                                       "Unbounded knapsack — count ways"),
    (7,  "Target Sum",                                 "Medium", 494,  "https://leetcode.com/problems/target-sum/",                                           "Assign +/- ; DP or DFS+memo"),
    (8,  "Maximal Square",                             "Medium", 221,  "https://leetcode.com/problems/maximal-square/",                                       "dp[i][j] = min of 3 neighbors + 1"),
    (9,  "Longest Increasing Path in Matrix",          "Hard",   329,  "https://leetcode.com/problems/longest-increasing-path-in-a-matrix/",                 "DFS + memo"),
    (10, "Distinct Subsequences",                      "Hard",   115,  "https://leetcode.com/problems/distinct-subsequences/",                               "Count ways to embed t in s"),
    (11, "Burst Balloons",                             "Hard",   312,  "https://leetcode.com/problems/burst-balloons/",                                       "Interval DP; last balloon"),
    (12, "Regular Expression Matching",                "Hard",   10,   "https://leetcode.com/problems/regular-expression-matching/",                          "2D DP; '*' = 0 or more"),
    (13, "Wildcard Matching",                          "Hard",   44,   "https://leetcode.com/problems/wildcard-matching/",                                    "'?' = any; '*' = any seq"),
    (14, "Best Time to Buy Sell — Cooldown",           "Medium", 309,  "https://leetcode.com/problems/best-time-to-buy-and-sell-stock-with-cooldown/",       "State machine DP"),
    (15, "Best Time to Buy Sell Stock III",            "Hard",   123,  "https://leetcode.com/problems/best-time-to-buy-and-sell-stock-iii/",                 "At most 2 transactions"),
    (16, "Minimum Difficulty of Job Schedule",         "Hard",   1335, "https://leetcode.com/problems/minimum-difficulty-of-a-job-schedule/",                "DP[day][idx]"),
    (17, "Stone Game",                                 "Medium", 877,  "https://leetcode.com/problems/stone-game/",                                           "Math insight or interval DP"),
    (18, "Super Egg Drop",                             "Hard",   887,  "https://leetcode.com/problems/super-egg-drop/",                                       "DP + binary search or math"),
]

TOPICS["Greedy"] = [
    (1,  "Gas Station",                          "Medium", 134,  "https://leetcode.com/problems/gas-station/",                             "Net gain; circular"),
    (2,  "Partition Labels",                      "Medium", 763,  "https://leetcode.com/problems/partition-labels/",                        "Last occurrence map"),
    (3,  "Valid Parenthesis String",              "Medium", 678,  "https://leetcode.com/problems/valid-parenthesis-string/",               "Track min/max open count"),
    (4,  "Minimum Arrows to Burst Balloons",      "Medium", 452,  "https://leetcode.com/problems/minimum-number-of-arrows-to-burst-balloons/", "Sort by end; greedy"),
    (5,  "Non-overlapping Intervals",             "Medium", 435,  "https://leetcode.com/problems/non-overlapping-intervals/",              "Sort by end; keep most"),
    (6,  "Merge Intervals",                       "Medium", 56,   "https://leetcode.com/problems/merge-intervals/",                         "Sort by start; merge"),
    (7,  "Insert Interval",                       "Medium", 57,   "https://leetcode.com/problems/insert-interval/",                         "Linear scan; merge overlaps"),
    (8,  "Candy",                                 "Hard",   135,  "https://leetcode.com/problems/candy/",                                    "Two passes L→R, R→L"),
    (9,  "Boats to Save People",                  "Medium", 881,  "https://leetcode.com/problems/boats-to-save-people/",                    "Sort + two pointers"),
    (10, "Task Scheduler",                        "Medium", 621,  "https://leetcode.com/problems/task-scheduler/",                          "Formula or greedy simulation"),
    (11, "Hand of Straights",                     "Medium", 846,  "https://leetcode.com/problems/hand-of-straights/",                      "Sort + greedy decrement"),
]

TOPICS["Intervals"] = [
    (1,  "Insert Interval",             "Medium", 57,   "https://leetcode.com/problems/insert-interval/",                            "Linear scan; merge overlaps"),
    (2,  "Merge Intervals",             "Medium", 56,   "https://leetcode.com/problems/merge-intervals/",                            "Sort by start; merge"),
    (3,  "Non-overlapping Intervals",   "Medium", 435,  "https://leetcode.com/problems/non-overlapping-intervals/",                  "Sort by end; count removed"),
    (4,  "Meeting Rooms",               "Easy",   252,  "https://leetcode.com/problems/meeting-rooms/",                              "Sort; check overlap"),
    (5,  "Meeting Rooms II",            "Medium", 253,  "https://leetcode.com/problems/meeting-rooms-ii/",                           "Min-heap of end times"),
    (6,  "Employee Free Time",          "Hard",   759,  "https://leetcode.com/problems/employee-free-time/",                         "Merge all; find gaps"),
    (7,  "Minimum Interval per Query",  "Hard",   1851, "https://leetcode.com/problems/minimum-interval-to-include-each-query/",     "Sort + min-heap"),
    (8,  "Interval List Intersections", "Medium", 986,  "https://leetcode.com/problems/interval-list-intersections/",               "Two pointers on both lists"),
]

TOPICS["Math & Geometry"] = [
    (1,  "Rotate Image",              "Medium", 48,   "https://leetcode.com/problems/rotate-image/",              "Transpose + reverse rows"),
    (2,  "Spiral Matrix",             "Medium", 54,   "https://leetcode.com/problems/spiral-matrix/",             "4-boundary simulation"),
    (3,  "Set Matrix Zeroes",         "Medium", 73,   "https://leetcode.com/problems/set-matrix-zeroes/",         "First row/col as flags"),
    (4,  "Happy Number",              "Easy",   202,  "https://leetcode.com/problems/happy-number/",              "Floyd cycle or hash set"),
    (5,  "Pow(x, n)",                 "Medium", 50,   "https://leetcode.com/problems/powx-n/",                    "Fast exponentiation"),
    (6,  "Multiply Strings",          "Medium", 43,   "https://leetcode.com/problems/multiply-strings/",          "Grade-school multiplication"),
    (7,  "Count Primes",              "Medium", 204,  "https://leetcode.com/problems/count-primes/",              "Sieve of Eratosthenes"),
    (8,  "Excel Sheet Column Number", "Easy",   171,  "https://leetcode.com/problems/excel-sheet-column-number/", "Base 26"),
    (9,  "Plus One",                  "Easy",   66,   "https://leetcode.com/problems/plus-one/",                  "Carry propagation"),
    (10, "Max Points on a Line",      "Hard",   149,  "https://leetcode.com/problems/max-points-on-a-line/",      "Slope hash map"),
    (11, "Integer to Roman",          "Medium", 12,   "https://leetcode.com/problems/integer-to-roman/",          "Greedy subtraction"),
    (12, "Roman to Integer",          "Easy",   13,   "https://leetcode.com/problems/roman-to-integer/",          "Previous value comparison"),
]

TOPICS["Bit Manipulation"] = [
    (1,  "Single Number",                    "Easy",   136, "https://leetcode.com/problems/single-number/",                           "XOR all"),
    (2,  "Number of 1 Bits",                 "Easy",   191, "https://leetcode.com/problems/number-of-1-bits/",                        "n & (n-1)"),
    (3,  "Counting Bits",                    "Easy",   338, "https://leetcode.com/problems/counting-bits/",                           "DP: dp[i] = dp[i>>1] + (i&1)"),
    (4,  "Reverse Bits",                     "Easy",   190, "https://leetcode.com/problems/reverse-bits/",                            "Shift and mask"),
    (5,  "Missing Number",                   "Easy",   268, "https://leetcode.com/problems/missing-number/",                          "XOR or Gauss sum"),
    (6,  "Sum of Two Integers",              "Medium", 371, "https://leetcode.com/problems/sum-of-two-integers/",                     "XOR + carry"),
    (7,  "Single Number II",                 "Medium", 137, "https://leetcode.com/problems/single-number-ii/",                        "Bit count mod 3"),
    (8,  "Single Number III",                "Medium", 260, "https://leetcode.com/problems/single-number-iii/",                       "XOR split by diff bit"),
    (9,  "Maximum XOR of Two Numbers",       "Medium", 421, "https://leetcode.com/problems/maximum-xor-of-two-numbers-in-an-array/",  "Bit trie or prefix"),
    (10, "Bitwise AND of Numbers Range",     "Medium", 201, "https://leetcode.com/problems/bitwise-and-of-numbers-range/",            "Common prefix"),
]

TOPICS["Design / OOP"] = [
    (1,  "LRU Cache",                          "Medium", 146,  "https://leetcode.com/problems/lru-cache/",                                 "Doubly-LL + hash map"),
    (2,  "LFU Cache",                          "Hard",   460,  "https://leetcode.com/problems/lfu-cache/",                                 "Two hash maps + freq tracking"),
    (3,  "Design HashMap",                     "Easy",   706,  "https://leetcode.com/problems/design-hashmap/",                            "Array of buckets"),
    (4,  "Design HashSet",                     "Easy",   705,  "https://leetcode.com/problems/design-hashset/",                            "Array of buckets"),
    (5,  "Design Twitter",                     "Medium", 355,  "https://leetcode.com/problems/design-twitter/",                            "Min-heap merge k lists"),
    (6,  "Design Underground System",          "Medium", 1396, "https://leetcode.com/problems/design-underground-system/",                 "Two hash maps"),
    (7,  "Snapshot Array",                     "Medium", 1146, "https://leetcode.com/problems/snapshot-array/",                            "Binary search on history"),
    (8,  "Implement Trie (Prefix Tree)",       "Medium", 208,  "https://leetcode.com/problems/implement-trie-prefix-tree/",               "TrieNode children dict"),
    (9,  "Range Sum Query — Mutable",          "Medium", 307,  "https://leetcode.com/problems/range-sum-query-mutable/",                   "Segment tree or BIT"),
    (10, "My Calendar I",                      "Medium", 729,  "https://leetcode.com/problems/my-calendar-i/",                            "Binary search on sorted list"),
    (11, "Flatten Nested List Iterator",       "Medium", 341,  "https://leetcode.com/problems/flatten-nested-list-iterator/",             "Stack-based flattening"),
    (12, "Serialize and Deserialize BST",      "Medium", 449,  "https://leetcode.com/problems/serialize-and-deserialize-bst/",            "Preorder + BST property"),
    (13, "Time Based Key-Value Store",         "Medium", 981,  "https://leetcode.com/problems/time-based-key-value-store/",               "Binary search on timestamps"),
    (14, "Find Median from Data Stream",       "Hard",   295,  "https://leetcode.com/problems/find-median-from-data-stream/",             "Two heaps (max + min)"),
]

TOPICS["String Manipulation"] = [
    (1,  "Longest Common Prefix",              "Easy",   14,  "https://leetcode.com/problems/longest-common-prefix/",                          "Vertical scan or binary search"),
    (2,  "Reverse Words in a String",          "Medium", 151, "https://leetcode.com/problems/reverse-words-in-a-string/",                     "Split + reverse"),
    (3,  "String to Integer (atoi)",           "Medium", 8,   "https://leetcode.com/problems/string-to-integer-atoi/",                        "Edge cases: spaces, sign, overflow"),
    (4,  "ZigZag Conversion",                  "Medium", 6,   "https://leetcode.com/problems/zigzag-conversion/",                             "Row simulation"),
    (5,  "Simplify Path",                      "Medium", 71,  "https://leetcode.com/problems/simplify-path/",                                 "Stack with split('/')"),
    (6,  "Text Justification",                 "Hard",   68,  "https://leetcode.com/problems/text-justification/",                            "Greedy space distribution"),
    (7,  "Minimum Window Substring",           "Hard",   76,  "https://leetcode.com/problems/minimum-window-substring/",                      "Sliding window + char count"),
    (8,  "Longest Valid Parentheses",          "Hard",   32,  "https://leetcode.com/problems/longest-valid-parentheses/",                     "Stack or DP"),
    (9,  "Minimum Window Subsequence",         "Hard",   727, "https://leetcode.com/problems/minimum-window-subsequence/",                    "Two-pointer forward+backward"),
    (10, "Count and Say",                      "Medium", 38,  "https://leetcode.com/problems/count-and-say/",                                 "Simulation"),
]


# ═══════════════════════════════════════════════════════════════════════════════
#  COMPANY DEFINITIONS
#  color: primary brand hex color (6 chars, no #)
#  tips:  list of interview tips
#  hf:    high-frequency questions (no, problem, difficulty, lc#, url, why)
# ═══════════════════════════════════════════════════════════════════════════════

COMPANIES = {}

# ─── Intuit ───────────────────────────────────────────────────────────────────
COMPANIES["Intuit"] = {
    "color": "0A66C2",
    "tips": [
        "Intuit values clarity and communication — explain your thought process as you code.",
        "Focus on practical coding: Intuit interviews heavily test Arrays, Strings, and DP.",
        "Be ready for system design questions around financial data processing and APIs.",
        "Expect behavioral questions around ownership, customer focus, and iterative delivery.",
        "Medium-difficulty problems dominate Intuit's coding rounds — master the NeetCode 150.",
    ],
    "hf": [
        (1,  "Two Sum",                                  "Easy",   1,    "https://leetcode.com/problems/two-sum/",                               "Hash map; most asked warmup"),
        (2,  "Valid Parentheses",                        "Easy",   20,   "https://leetcode.com/problems/valid-parentheses/",                     "Stack; common screening Q"),
        (3,  "Longest Substring Without Repeating",      "Medium", 3,    "https://leetcode.com/problems/longest-substring-without-repeating-characters/", "Sliding window pattern"),
        (4,  "Add Two Numbers",                          "Medium", 2,    "https://leetcode.com/problems/add-two-numbers/",                       "LL carry propagation"),
        (5,  "Merge Two Sorted Lists",                   "Easy",   21,   "https://leetcode.com/problems/merge-two-sorted-lists/",                "Dummy head merge"),
        (6,  "Maximum Subarray",                         "Medium", 53,   "https://leetcode.com/problems/maximum-subarray/",                      "Kadane's algorithm"),
        (7,  "Climbing Stairs",                          "Easy",   70,   "https://leetcode.com/problems/climbing-stairs/",                       "Fibonacci DP"),
        (8,  "Best Time to Buy and Sell Stock",          "Easy",   121,  "https://leetcode.com/problems/best-time-to-buy-and-sell-stock/",       "Track min; greedy"),
        (9,  "Number of Islands",                        "Medium", 200,  "https://leetcode.com/problems/number-of-islands/",                     "BFS/DFS flood-fill"),
        (10, "Course Schedule",                          "Medium", 207,  "https://leetcode.com/problems/course-schedule/",                       "Topological sort / cycle"),
        (11, "Word Break",                               "Medium", 139,  "https://leetcode.com/problems/word-break/",                            "1D DP"),
        (12, "Coin Change",                              "Medium", 322,  "https://leetcode.com/problems/coin-change/",                           "Unbounded knapsack DP"),
        (13, "House Robber",                             "Medium", 198,  "https://leetcode.com/problems/house-robber/",                          "1D DP; no adjacent"),
        (14, "Longest Palindromic Substring",            "Medium", 5,    "https://leetcode.com/problems/longest-palindromic-substring/",         "Expand around center"),
        (15, "3Sum",                                     "Medium", 15,   "https://leetcode.com/problems/3sum/",                                  "Sort + two pointers"),
        (16, "Decode Ways",                              "Medium", 91,   "https://leetcode.com/problems/decode-ways/",                           "1D DP; leading zeros"),
        (17, "Unique Paths",                             "Medium", 62,   "https://leetcode.com/problems/unique-paths/",                          "Grid DP"),
        (18, "Jump Game",                                "Medium", 55,   "https://leetcode.com/problems/jump-game/",                             "Greedy max reach"),
        (19, "Word Search",                              "Medium", 79,   "https://leetcode.com/problems/word-search/",                           "DFS + visited"),
        (20, "Merge Intervals",                          "Medium", 56,   "https://leetcode.com/problems/merge-intervals/",                       "Sort + merge"),
        (21, "LRU Cache",                                "Medium", 146,  "https://leetcode.com/problems/lru-cache/",                             "Doubly-LL + hash map"),
        (22, "Product of Array Except Self",             "Medium", 238,  "https://leetcode.com/problems/product-of-array-except-self/",          "Prefix & suffix"),
        (23, "Top K Frequent Elements",                  "Medium", 347,  "https://leetcode.com/problems/top-k-frequent-elements/",               "Heap or bucket sort"),
        (24, "Binary Tree Level Order Traversal",        "Medium", 102,  "https://leetcode.com/problems/binary-tree-level-order-traversal/",     "BFS with queue"),
        (25, "Validate Binary Search Tree",              "Medium", 98,   "https://leetcode.com/problems/validate-binary-search-tree/",           "Min/max range DFS"),
    ],
}

# ─── Adobe ───────────────────────────────────────────────────────────────────
COMPANIES["Adobe"] = {
    "color": "FF0000",
    "tips": [
        "Adobe interviews focus on arrays, strings, trees, and DP — expect 2–3 coding rounds.",
        "Emphasis on clean, production-quality code. Explain trade-offs clearly.",
        "System design rounds test knowledge of media processing, cloud APIs, and content pipelines.",
        "Matrix/grid problems (Spiral Matrix, Rotate Image) are Adobe favourites.",
        "Behavioural: ownership, cross-functional collaboration, and delivering creative solutions.",
    ],
    "hf": [
        (1,  "Two Sum",                                  "Easy",   1,    "https://leetcode.com/problems/two-sum/",                               "Warmup; hash map"),
        (2,  "Merge Intervals",                          "Medium", 56,   "https://leetcode.com/problems/merge-intervals/",                       "Sort + merge"),
        (3,  "Number of Islands",                        "Medium", 200,  "https://leetcode.com/problems/number-of-islands/",                     "BFS/DFS flood-fill"),
        (4,  "Longest Palindromic Substring",            "Medium", 5,    "https://leetcode.com/problems/longest-palindromic-substring/",         "Expand around center"),
        (5,  "Word Search",                              "Medium", 79,   "https://leetcode.com/problems/word-search/",                           "DFS + visited"),
        (6,  "Spiral Matrix",                            "Medium", 54,   "https://leetcode.com/problems/spiral-matrix/",                         "4-boundary simulation"),
        (7,  "Valid Parentheses",                        "Easy",   20,   "https://leetcode.com/problems/valid-parentheses/",                     "Stack"),
        (8,  "Rotate Image",                             "Medium", 48,   "https://leetcode.com/problems/rotate-image/",                          "Transpose + reverse rows"),
        (9,  "Set Matrix Zeroes",                        "Medium", 73,   "https://leetcode.com/problems/set-matrix-zeroes/",                     "In-place flag first row/col"),
        (10, "Maximum Subarray",                         "Medium", 53,   "https://leetcode.com/problems/maximum-subarray/",                      "Kadane's"),
        (11, "Trapping Rain Water",                      "Hard",   42,   "https://leetcode.com/problems/trapping-rain-water/",                   "Two pointers"),
        (12, "LRU Cache",                                "Medium", 146,  "https://leetcode.com/problems/lru-cache/",                             "Doubly-LL + hash map"),
        (13, "Course Schedule",                          "Medium", 207,  "https://leetcode.com/problems/course-schedule/",                       "Topo sort"),
        (14, "Median of Two Sorted Arrays",              "Hard",   4,    "https://leetcode.com/problems/median-of-two-sorted-arrays/",           "Binary search on partition"),
        (15, "Edit Distance",                            "Medium", 72,   "https://leetcode.com/problems/edit-distance/",                         "2D DP"),
        (16, "Word Ladder",                              "Hard",   127,  "https://leetcode.com/problems/word-ladder/",                           "BFS with word transforms"),
        (17, "Longest Common Subsequence",               "Medium", 1143, "https://leetcode.com/problems/longest-common-subsequence/",            "Classic 2D DP"),
        (18, "Regular Expression Matching",              "Hard",   10,   "https://leetcode.com/problems/regular-expression-matching/",           "2D DP; '*' = 0 or more"),
        (19, "Serialize and Deserialize Binary Tree",    "Hard",   297,  "https://leetcode.com/problems/serialize-and-deserialize-binary-tree/", "BFS or preorder"),
        (20, "Jump Game II",                             "Medium", 45,   "https://leetcode.com/problems/jump-game-ii/",                          "Greedy levels"),
        (21, "Binary Tree Maximum Path Sum",             "Hard",   124,  "https://leetcode.com/problems/binary-tree-maximum-path-sum/",          "Post-order; global max"),
        (22, "Decode Ways",                              "Medium", 91,   "https://leetcode.com/problems/decode-ways/",                           "1D DP"),
        (23, "Kth Largest Element in Array",             "Medium", 215,  "https://leetcode.com/problems/kth-largest-element-in-an-array/",       "Quick select or heap"),
        (24, "Find Median from Data Stream",             "Hard",   295,  "https://leetcode.com/problems/find-median-from-data-stream/",          "Two heaps"),
        (25, "Clone Graph",                              "Medium", 133,  "https://leetcode.com/problems/clone-graph/",                           "BFS/DFS + hash map"),
    ],
}

# ─── Microsoft ───────────────────────────────────────────────────────────────
COMPANIES["Microsoft"] = {
    "color": "0078D4",
    "tips": [
        "Microsoft screens for problem-solving clarity. Always walk through examples before coding.",
        "Expect 4–5 rounds: coding, design, behavioural. Growth mindset is highly valued.",
        "Trees, Graphs, DP, and Design are the top four topic areas at Microsoft.",
        "Clean code over clever one-liners. Readability and maintainability matter.",
        "Know Azure-scale system design: distributed locks, event streaming, caching patterns.",
    ],
    "hf": [
        (1,  "Reverse Linked List",                      "Easy",   206,  "https://leetcode.com/problems/reverse-linked-list/",                   "Iterative & recursive"),
        (2,  "LRU Cache",                                "Medium", 146,  "https://leetcode.com/problems/lru-cache/",                             "Doubly-LL + hash map"),
        (3,  "Clone Graph",                              "Medium", 133,  "https://leetcode.com/problems/clone-graph/",                           "BFS/DFS + hash map"),
        (4,  "Copy List with Random Pointer",            "Medium", 138,  "https://leetcode.com/problems/copy-list-with-random-pointer/",         "Hash map or weave"),
        (5,  "Word Search",                              "Medium", 79,   "https://leetcode.com/problems/word-search/",                           "DFS + visited"),
        (6,  "Binary Tree Level Order Traversal",        "Medium", 102,  "https://leetcode.com/problems/binary-tree-level-order-traversal/",     "BFS with queue"),
        (7,  "Number of Islands",                        "Medium", 200,  "https://leetcode.com/problems/number-of-islands/",                     "BFS/DFS flood-fill"),
        (8,  "Longest Palindromic Substring",            "Medium", 5,    "https://leetcode.com/problems/longest-palindromic-substring/",         "Expand around center"),
        (9,  "Kth Largest Element in Array",             "Medium", 215,  "https://leetcode.com/problems/kth-largest-element-in-an-array/",       "Quick select or heap"),
        (10, "Maximum Subarray",                         "Medium", 53,   "https://leetcode.com/problems/maximum-subarray/",                      "Kadane's"),
        (11, "Valid Parentheses",                        "Easy",   20,   "https://leetcode.com/problems/valid-parentheses/",                     "Stack"),
        (12, "Sort Colors",                              "Medium", 75,   "https://leetcode.com/problems/sort-colors/",                           "Dutch flag 3-pointer"),
        (13, "Maximum Depth of Binary Tree",             "Easy",   104,  "https://leetcode.com/problems/maximum-depth-of-binary-tree/",          "DFS"),
        (14, "Flatten Binary Tree to Linked List",       "Medium", 114,  "https://leetcode.com/problems/flatten-binary-tree-to-linked-list/",    "Morris traversal"),
        (15, "Serialize and Deserialize Binary Tree",    "Hard",   297,  "https://leetcode.com/problems/serialize-and-deserialize-binary-tree/", "Preorder + queue"),
        (16, "Count of Smaller Numbers After Self",      "Hard",   315,  "https://leetcode.com/problems/count-of-smaller-numbers-after-self/",   "Merge sort / BIT"),
        (17, "Validate Binary Search Tree",              "Medium", 98,   "https://leetcode.com/problems/validate-binary-search-tree/",           "Min/max DFS"),
        (18, "Course Schedule",                          "Medium", 207,  "https://leetcode.com/problems/course-schedule/",                       "Topo sort"),
        (19, "Word Break",                               "Medium", 139,  "https://leetcode.com/problems/word-break/",                            "1D DP"),
        (20, "Reverse String",                           "Easy",   344,  "https://leetcode.com/problems/reverse-string/",                        "Two pointers"),
        (21, "Excel Sheet Column Number",                "Easy",   171,  "https://leetcode.com/problems/excel-sheet-column-number/",             "Base 26"),
        (22, "Spiral Matrix",                            "Medium", 54,   "https://leetcode.com/problems/spiral-matrix/",                         "4-boundary simulation"),
        (23, "Two Sum",                                  "Easy",   1,    "https://leetcode.com/problems/two-sum/",                               "Hash map"),
        (24, "Merge K Sorted Lists",                     "Hard",   23,   "https://leetcode.com/problems/merge-k-sorted-lists/",                  "Min-heap"),
        (25, "Design Underground System",                "Medium", 1396, "https://leetcode.com/problems/design-underground-system/",             "Two hash maps"),
    ],
}

# ─── Amazon ───────────────────────────────────────────────────────────────────
COMPANIES["Amazon"] = {
    "color": "232F3E",
    "tips": [
        "Amazon's Leadership Principles (LPs) are assessed in every round — prepare STAR stories for each LP.",
        "Coding rounds test Trees, Graphs, Sliding Window, Heaps, and DP heavily.",
        "Scale-thinking matters: always discuss your solution's time/space complexity at Amazon scale.",
        "Two coding rounds per loop; the bar-raiser round is usually harder — practice Hard problems.",
        "OOP design and system design (distributed systems, event-driven) are common in senior rounds.",
    ],
    "hf": [
        (1,  "Two Sum",                                  "Easy",   1,    "https://leetcode.com/problems/two-sum/",                               "Hash map"),
        (2,  "LRU Cache",                                "Medium", 146,  "https://leetcode.com/problems/lru-cache/",                             "Amazon top-asked design Q"),
        (3,  "Number of Islands",                        "Medium", 200,  "https://leetcode.com/problems/number-of-islands/",                     "BFS/DFS flood-fill"),
        (4,  "Trapping Rain Water",                      "Hard",   42,   "https://leetcode.com/problems/trapping-rain-water/",                   "Two pointers"),
        (5,  "Word Ladder",                              "Hard",   127,  "https://leetcode.com/problems/word-ladder/",                           "BFS transformation graph"),
        (6,  "Top K Frequent Elements",                  "Medium", 347,  "https://leetcode.com/problems/top-k-frequent-elements/",               "Heap or bucket sort"),
        (7,  "Merge Intervals",                          "Medium", 56,   "https://leetcode.com/problems/merge-intervals/",                       "Sort + merge"),
        (8,  "Meeting Rooms II",                         "Medium", 253,  "https://leetcode.com/problems/meeting-rooms-ii/",                      "Min-heap end times"),
        (9,  "Course Schedule",                          "Medium", 207,  "https://leetcode.com/problems/course-schedule/",                       "Topo sort"),
        (10, "K Closest Points to Origin",               "Medium", 973,  "https://leetcode.com/problems/k-closest-points-to-origin/",            "Max-heap size k"),
        (11, "Sliding Window Maximum",                   "Hard",   239,  "https://leetcode.com/problems/sliding-window-maximum/",                "Monotonic deque"),
        (12, "Word Break II",                            "Hard",   140,  "https://leetcode.com/problems/word-break-ii/",                         "Backtracking + memo"),
        (13, "Copy List with Random Pointer",            "Medium", 138,  "https://leetcode.com/problems/copy-list-with-random-pointer/",         "Hash map or weave"),
        (14, "Maximum Profit Job Scheduling",            "Hard",   1235, "https://leetcode.com/problems/maximum-profit-in-job-scheduling/",      "Sort + DP + binary search"),
        (15, "Find All Anagrams in String",              "Medium", 438,  "https://leetcode.com/problems/find-all-anagrams-in-a-string/",         "Fixed sliding window"),
        (16, "Fruit Into Baskets",                       "Medium", 904,  "https://leetcode.com/problems/fruit-into-baskets/",                    "At most 2 distinct"),
        (17, "Reorder Data in Log Files",                "Medium", 937,  "https://leetcode.com/problems/reorder-data-in-log-files/",             "Custom sort; Amazon exclusive"),
        (18, "Minimum Difficulty of Job Schedule",       "Hard",   1335, "https://leetcode.com/problems/minimum-difficulty-of-a-job-schedule/",  "DP[day][idx] — Amazon classic"),
        (19, "Critical Connections in Network",          "Hard",   1192, "https://leetcode.com/problems/critical-connections-in-a-network/",     "Tarjan's bridge"),
        (20, "Merge K Sorted Lists",                     "Hard",   23,   "https://leetcode.com/problems/merge-k-sorted-lists/",                  "Min-heap"),
        (21, "Design Underground System",                "Medium", 1396, "https://leetcode.com/problems/design-underground-system/",             "Two hash maps"),
        (22, "Maximum Subarray",                         "Medium", 53,   "https://leetcode.com/problems/maximum-subarray/",                      "Kadane's"),
        (23, "Coin Change",                              "Medium", 322,  "https://leetcode.com/problems/coin-change/",                           "Unbounded knapsack DP"),
        (24, "Subsets",                                  "Medium", 78,   "https://leetcode.com/problems/subsets/",                               "Backtracking"),
        (25, "Partition Labels",                         "Medium", 763,  "https://leetcode.com/problems/partition-labels/",                      "Last occurrence greedy"),
    ],
}

# ─── JP Morgan ────────────────────────────────────────────────────────────────
COMPANIES["JP Morgan"] = {
    "color": "003087",
    "tips": [
        "JP Morgan interviews focus on core DSA fundamentals — expect clear, optimised solutions.",
        "Financial context matters: know time-series, sliding window, and interval problems well.",
        "Coding rounds are typically 2 rounds; system design is expected for senior engineers.",
        "Be ready for SQL questions and data-structure design in quant-adjacent roles.",
        "Communication and financial domain awareness are valued alongside algorithmic skill.",
    ],
    "hf": [
        (1,  "Two Sum",                                  "Easy",   1,    "https://leetcode.com/problems/two-sum/",                               "Hash map"),
        (2,  "Valid Parentheses",                        "Easy",   20,   "https://leetcode.com/problems/valid-parentheses/",                     "Stack"),
        (3,  "Longest Substring Without Repeating",      "Medium", 3,    "https://leetcode.com/problems/longest-substring-without-repeating-characters/", "Sliding window"),
        (4,  "Merge Two Sorted Lists",                   "Easy",   21,   "https://leetcode.com/problems/merge-two-sorted-lists/",                "Dummy head"),
        (5,  "Add Two Numbers",                          "Medium", 2,    "https://leetcode.com/problems/add-two-numbers/",                       "Carry propagation"),
        (6,  "Maximum Subarray",                         "Medium", 53,   "https://leetcode.com/problems/maximum-subarray/",                      "Kadane's"),
        (7,  "Climbing Stairs",                          "Easy",   70,   "https://leetcode.com/problems/climbing-stairs/",                       "Fibonacci DP"),
        (8,  "Binary Tree Level Order Traversal",        "Medium", 102,  "https://leetcode.com/problems/binary-tree-level-order-traversal/",     "BFS"),
        (9,  "Number of Islands",                        "Medium", 200,  "https://leetcode.com/problems/number-of-islands/",                     "BFS/DFS"),
        (10, "Coin Change",                              "Medium", 322,  "https://leetcode.com/problems/coin-change/",                           "DP — financial problem flavour"),
        (11, "House Robber",                             "Medium", 198,  "https://leetcode.com/problems/house-robber/",                          "1D DP"),
        (12, "Course Schedule",                          "Medium", 207,  "https://leetcode.com/problems/course-schedule/",                       "Topo sort"),
        (13, "Product of Array Except Self",             "Medium", 238,  "https://leetcode.com/problems/product-of-array-except-self/",          "Prefix & suffix"),
        (14, "Find Minimum in Rotated Sorted Array",     "Medium", 153,  "https://leetcode.com/problems/find-minimum-in-rotated-sorted-array/",  "Binary search"),
        (15, "Word Break",                               "Medium", 139,  "https://leetcode.com/problems/word-break/",                            "1D DP"),
        (16, "Decode Ways",                              "Medium", 91,   "https://leetcode.com/problems/decode-ways/",                           "1D DP"),
        (17, "Maximum Product Subarray",                 "Medium", 152,  "https://leetcode.com/problems/maximum-product-subarray/",              "Track min & max"),
        (18, "Merge Intervals",                          "Medium", 56,   "https://leetcode.com/problems/merge-intervals/",                       "Sort + merge"),
        (19, "Meeting Rooms II",                         "Medium", 253,  "https://leetcode.com/problems/meeting-rooms-ii/",                      "Min-heap end times"),
        (20, "Top K Frequent Elements",                  "Medium", 347,  "https://leetcode.com/problems/top-k-frequent-elements/",               "Heap"),
        (21, "3Sum",                                     "Medium", 15,   "https://leetcode.com/problems/3sum/",                                  "Sort + two pointers"),
        (22, "Kth Largest Element in Array",             "Medium", 215,  "https://leetcode.com/problems/kth-largest-element-in-an-array/",       "Quick select / heap"),
        (23, "Best Time to Buy and Sell Stock",          "Easy",   121,  "https://leetcode.com/problems/best-time-to-buy-and-sell-stock/",       "Financial context"),
        (24, "Best Time to Buy and Sell Stock III",      "Hard",   123,  "https://leetcode.com/problems/best-time-to-buy-and-sell-stock-iii/",   "At most 2 transactions"),
        (25, "LRU Cache",                                "Medium", 146,  "https://leetcode.com/problems/lru-cache/",                             "Design Q"),
    ],
}

# ─── Morgan Stanley ──────────────────────────────────────────────────────────
COMPANIES["Morgan Stanley"] = {
    "color": "1A1A2E",
    "tips": [
        "Morgan Stanley values strong fundamentals — focus on arrays, strings, trees, and DP.",
        "Financial algorithms (stock prices, portfolio optimisation) appear as contextual DP/greedy.",
        "Coding is paired with strong system design expectations in senior rounds.",
        "Communicate trade-offs: latency vs. throughput, memory vs. time.",
        "Expect questions about data structures used in real trading systems.",
    ],
    "hf": [
        (1,  "Two Sum",                                  "Easy",   1,    "https://leetcode.com/problems/two-sum/",                               "Hash map"),
        (2,  "Valid Parentheses",                        "Easy",   20,   "https://leetcode.com/problems/valid-parentheses/",                     "Stack"),
        (3,  "Reverse Linked List",                      "Easy",   206,  "https://leetcode.com/problems/reverse-linked-list/",                   "Iterative"),
        (4,  "Best Time to Buy and Sell Stock",          "Easy",   121,  "https://leetcode.com/problems/best-time-to-buy-and-sell-stock/",       "Financial context; greedy"),
        (5,  "Merge Two Sorted Lists",                   "Easy",   21,   "https://leetcode.com/problems/merge-two-sorted-lists/",                "Dummy head"),
        (6,  "Maximum Subarray",                         "Medium", 53,   "https://leetcode.com/problems/maximum-subarray/",                      "Kadane's"),
        (7,  "Add Two Numbers",                          "Medium", 2,    "https://leetcode.com/problems/add-two-numbers/",                       "Carry propagation"),
        (8,  "Longest Substring Without Repeating",      "Medium", 3,    "https://leetcode.com/problems/longest-substring-without-repeating-characters/", "Sliding window"),
        (9,  "3Sum",                                     "Medium", 15,   "https://leetcode.com/problems/3sum/",                                  "Sort + two pointers"),
        (10, "Word Break",                               "Medium", 139,  "https://leetcode.com/problems/word-break/",                            "1D DP"),
        (11, "Coin Change",                              "Medium", 322,  "https://leetcode.com/problems/coin-change/",                           "Unbounded knapsack"),
        (12, "LRU Cache",                                "Medium", 146,  "https://leetcode.com/problems/lru-cache/",                             "Design"),
        (13, "Course Schedule",                          "Medium", 207,  "https://leetcode.com/problems/course-schedule/",                       "Topo sort"),
        (14, "Number of Islands",                        "Medium", 200,  "https://leetcode.com/problems/number-of-islands/",                     "BFS/DFS"),
        (15, "Binary Tree Level Order Traversal",        "Medium", 102,  "https://leetcode.com/problems/binary-tree-level-order-traversal/",     "BFS"),
        (16, "Decode Ways",                              "Medium", 91,   "https://leetcode.com/problems/decode-ways/",                           "1D DP"),
        (17, "Maximum Product Subarray",                 "Medium", 152,  "https://leetcode.com/problems/maximum-product-subarray/",              "Track min & max"),
        (18, "Jump Game",                                "Medium", 55,   "https://leetcode.com/problems/jump-game/",                             "Greedy max reach"),
        (19, "Trapping Rain Water",                      "Hard",   42,   "https://leetcode.com/problems/trapping-rain-water/",                   "Two pointers"),
        (20, "Median of Two Sorted Arrays",              "Hard",   4,    "https://leetcode.com/problems/median-of-two-sorted-arrays/",           "Binary search on partition"),
        (21, "Find Median from Data Stream",             "Hard",   295,  "https://leetcode.com/problems/find-median-from-data-stream/",          "Two heaps — trading systems"),
        (22, "Sliding Window Maximum",                   "Hard",   239,  "https://leetcode.com/problems/sliding-window-maximum/",                "Monotonic deque"),
        (23, "Best Time to Buy Sell Stock III",          "Hard",   123,  "https://leetcode.com/problems/best-time-to-buy-and-sell-stock-iii/",   "At most 2 transactions"),
        (24, "Merge Intervals",                          "Medium", 56,   "https://leetcode.com/problems/merge-intervals/",                       "Sort + merge"),
        (25, "Longest Common Subsequence",               "Medium", 1143, "https://leetcode.com/problems/longest-common-subsequence/",            "2D DP"),
    ],
}

# ─── Nvidia ───────────────────────────────────────────────────────────────────
COMPANIES["Nvidia"] = {
    "color": "76B900",
    "tips": [
        "Nvidia values deep algorithmic thinking — GPU-friendly data structures and parallel patterns.",
        "Expect matrix/grid problems (2D DP, spiral, BFS on grids) frequently.",
        "System design may touch on GPU memory hierarchy, parallel compute pipelines.",
        "Focus on DP, Graphs, and simulation-style problems.",
        "C++ proficiency is a strong plus; discuss memory and cache efficiency.",
    ],
    "hf": [
        (1,  "Two Sum",                                  "Easy",   1,    "https://leetcode.com/problems/two-sum/",                               "Hash map"),
        (2,  "Maximum Subarray",                         "Medium", 53,   "https://leetcode.com/problems/maximum-subarray/",                      "Kadane's"),
        (3,  "Climbing Stairs",                          "Easy",   70,   "https://leetcode.com/problems/climbing-stairs/",                       "Fibonacci DP"),
        (4,  "Reverse Linked List",                      "Easy",   206,  "https://leetcode.com/problems/reverse-linked-list/",                   "Iterative"),
        (5,  "Number of Islands",                        "Medium", 200,  "https://leetcode.com/problems/number-of-islands/",                     "BFS/DFS"),
        (6,  "Word Break",                               "Medium", 139,  "https://leetcode.com/problems/word-break/",                            "1D DP"),
        (7,  "Course Schedule",                          "Medium", 207,  "https://leetcode.com/problems/course-schedule/",                       "Topo sort"),
        (8,  "Longest Increasing Subsequence",           "Medium", 300,  "https://leetcode.com/problems/longest-increasing-subsequence/",        "DP O(n²) or patience sort"),
        (9,  "Clone Graph",                              "Medium", 133,  "https://leetcode.com/problems/clone-graph/",                           "BFS + hash map"),
        (10, "Decode Ways",                              "Medium", 91,   "https://leetcode.com/problems/decode-ways/",                           "1D DP"),
        (11, "Maximum Product Subarray",                 "Medium", 152,  "https://leetcode.com/problems/maximum-product-subarray/",              "Track min & max"),
        (12, "Minimum Path Sum",                         "Medium", 64,   "https://leetcode.com/problems/minimum-path-sum/",                      "Grid DP — GPU-grid analogy"),
        (13, "Coin Change",                              "Medium", 322,  "https://leetcode.com/problems/coin-change/",                           "Unbounded knapsack"),
        (14, "Spiral Matrix",                            "Medium", 54,   "https://leetcode.com/problems/spiral-matrix/",                         "Grid simulation"),
        (15, "Trapping Rain Water",                      "Hard",   42,   "https://leetcode.com/problems/trapping-rain-water/",                   "Two pointers"),
        (16, "Median of Two Sorted Arrays",              "Hard",   4,    "https://leetcode.com/problems/median-of-two-sorted-arrays/",           "Binary search on partition"),
        (17, "Edit Distance",                            "Medium", 72,   "https://leetcode.com/problems/edit-distance/",                         "2D DP"),
        (18, "N-Queens",                                 "Hard",   51,   "https://leetcode.com/problems/n-queens/",                              "Backtracking; constraint satisfaction"),
        (19, "Longest Common Subsequence",               "Medium", 1143, "https://leetcode.com/problems/longest-common-subsequence/",            "2D DP"),
        (20, "Matrix Diagonal Sum",                      "Easy",   1572, "https://leetcode.com/problems/matrix-diagonal-sum/",                   "Grid math — GPU-grid"),
        (21, "Unique Paths",                             "Medium", 62,   "https://leetcode.com/problems/unique-paths/",                          "Grid DP"),
        (22, "Maximal Square",                           "Medium", 221,  "https://leetcode.com/problems/maximal-square/",                        "2D DP; min of 3 neighbours"),
        (23, "Largest Rectangle in Histogram",           "Hard",   84,   "https://leetcode.com/problems/largest-rectangle-in-histogram/",        "Monotonic stack"),
        (24, "Pacific Atlantic Water Flow",              "Medium", 417,  "https://leetcode.com/problems/pacific-atlantic-water-flow/",           "Reverse DFS from coasts"),
        (25, "Swim in Rising Water",                     "Hard",   778,  "https://leetcode.com/problems/swim-in-rising-water/",                  "Dijkstra / binary search + BFS"),
    ],
}

# ─── Rubrik ───────────────────────────────────────────────────────────────────
COMPANIES["Rubrik"] = {
    "color": "00A1D9",
    "tips": [
        "Rubrik's interviews emphasise storage systems, data management, and distributed computing.",
        "Design questions often involve file systems, snapshots, and backup/recovery pipelines.",
        "Coding rounds focus on graphs, design, intervals, and DP.",
        "Expect at least one system design round at the SDE-2+ level.",
        "Strong fundamentals in concurrent data structures and OS concepts are a plus.",
    ],
    "hf": [
        (1,  "Two Sum",                                  "Easy",   1,    "https://leetcode.com/problems/two-sum/",                               "Hash map"),
        (2,  "Maximum Subarray",                         "Medium", 53,   "https://leetcode.com/problems/maximum-subarray/",                      "Kadane's"),
        (3,  "LRU Cache",                                "Medium", 146,  "https://leetcode.com/problems/lru-cache/",                             "Storage eviction policy"),
        (4,  "Number of Islands",                        "Medium", 200,  "https://leetcode.com/problems/number-of-islands/",                     "BFS/DFS"),
        (5,  "Course Schedule",                          "Medium", 207,  "https://leetcode.com/problems/course-schedule/",                       "Topo sort; dependency graph"),
        (6,  "Word Break",                               "Medium", 139,  "https://leetcode.com/problems/word-break/",                            "1D DP"),
        (7,  "Clone Graph",                              "Medium", 133,  "https://leetcode.com/problems/clone-graph/",                           "BFS + hash map"),
        (8,  "Meeting Rooms II",                         "Medium", 253,  "https://leetcode.com/problems/meeting-rooms-ii/",                      "Min-heap — scheduling"),
        (9,  "Merge Intervals",                          "Medium", 56,   "https://leetcode.com/problems/merge-intervals/",                       "Sort + merge"),
        (10, "Design File System",                       "Medium", 1166, "https://leetcode.com/problems/design-file-system/",                    "Trie or hash map — Rubrik core"),
        (11, "Minimum Cost to Connect Sticks",           "Medium", 1167, "https://leetcode.com/problems/minimum-cost-to-connect-sticks/",        "Min-heap"),
        (12, "Task Scheduler",                           "Medium", 621,  "https://leetcode.com/problems/task-scheduler/",                        "Greedy + cooldown"),
        (13, "K Closest Points to Origin",               "Medium", 973,  "https://leetcode.com/problems/k-closest-points-to-origin/",            "Max-heap size k"),
        (14, "Find Median from Data Stream",             "Hard",   295,  "https://leetcode.com/problems/find-median-from-data-stream/",          "Two heaps"),
        (15, "Serialize Deserialize Binary Tree",        "Hard",   297,  "https://leetcode.com/problems/serialize-and-deserialize-binary-tree/", "Snapshot/restore pattern"),
        (16, "Sliding Window Maximum",                   "Hard",   239,  "https://leetcode.com/problems/sliding-window-maximum/",                "Monotonic deque"),
        (17, "Word Ladder",                              "Hard",   127,  "https://leetcode.com/problems/word-ladder/",                           "BFS graph"),
        (18, "Critical Connections in Network",          "Hard",   1192, "https://leetcode.com/problems/critical-connections-in-a-network/",     "Tarjan's bridge — network infra"),
        (19, "Maximum Profit Job Scheduling",            "Hard",   1235, "https://leetcode.com/problems/maximum-profit-in-job-scheduling/",      "DP + binary search"),
        (20, "Snapshot Array",                           "Medium", 1146, "https://leetcode.com/problems/snapshot-array/",                        "Binary search on history — Rubrik classic"),
        (21, "LFU Cache",                                "Hard",   460,  "https://leetcode.com/problems/lfu-cache/",                             "Frequency-based eviction"),
        (22, "Design In-Memory File System",             "Hard",   588,  "https://leetcode.com/problems/design-in-memory-file-system/",          "Trie OOP — Rubrik very frequent"),
        (23, "Min Cost to Connect All Points",           "Medium", 1584, "https://leetcode.com/problems/min-cost-to-connect-all-points/",        "Prim's MST"),
        (24, "Coin Change",                              "Medium", 322,  "https://leetcode.com/problems/coin-change/",                           "DP"),
        (25, "Longest Increasing Subsequence",           "Medium", 300,  "https://leetcode.com/problems/longest-increasing-subsequence/",        "DP"),
    ],
}

# ─── Databricks ───────────────────────────────────────────────────────────────
COMPANIES["Databricks"] = {
    "color": "FF3621",
    "tips": [
        "Databricks interviews are demanding — expect Hard-level graph and DP problems.",
        "Distributed systems and Spark internals are tested in design rounds.",
        "Graphs, design patterns (cache, streaming), and interval scheduling are high-frequency.",
        "Strong Python/Scala coding is expected alongside algorithmic proficiency.",
        "Know CAP theorem, data lake patterns, and columnar storage internals.",
    ],
    "hf": [
        (1,  "Two Sum",                                  "Easy",   1,    "https://leetcode.com/problems/two-sum/",                               "Hash map"),
        (2,  "Maximum Subarray",                         "Medium", 53,   "https://leetcode.com/problems/maximum-subarray/",                      "Kadane's"),
        (3,  "Number of Islands",                        "Medium", 200,  "https://leetcode.com/problems/number-of-islands/",                     "BFS/DFS"),
        (4,  "Course Schedule",                          "Medium", 207,  "https://leetcode.com/problems/course-schedule/",                       "Topo sort"),
        (5,  "Clone Graph",                              "Medium", 133,  "https://leetcode.com/problems/clone-graph/",                           "BFS + hash map"),
        (6,  "Word Break",                               "Medium", 139,  "https://leetcode.com/problems/word-break/",                            "1D DP"),
        (7,  "Accounts Merge",                           "Medium", 721,  "https://leetcode.com/problems/accounts-merge/",                        "Union-Find — data dedup"),
        (8,  "Merge Intervals",                          "Medium", 56,   "https://leetcode.com/problems/merge-intervals/",                       "Sort + merge"),
        (9,  "Evaluate Division",                        "Medium", 399,  "https://leetcode.com/problems/evaluate-division/",                     "Weighted graph BFS"),
        (10, "Find Median from Data Stream",             "Hard",   295,  "https://leetcode.com/problems/find-median-from-data-stream/",          "Two heaps — streaming"),
        (11, "Alien Dictionary",                         "Hard",   269,  "https://leetcode.com/problems/alien-dictionary/",                      "Topo sort from char order"),
        (12, "Sliding Window Maximum",                   "Hard",   239,  "https://leetcode.com/problems/sliding-window-maximum/",                "Monotonic deque — windowed agg"),
        (13, "Word Ladder",                              "Hard",   127,  "https://leetcode.com/problems/word-ladder/",                           "BFS"),
        (14, "Critical Connections in Network",          "Hard",   1192, "https://leetcode.com/problems/critical-connections-in-a-network/",     "Tarjan's bridge"),
        (15, "Min Vertices to Reach All Nodes",          "Medium", 1557, "https://leetcode.com/problems/minimum-number-of-vertices-to-reach-all-nodes/", "Zero in-degree"),
        (16, "Longest Increasing Subsequence",           "Medium", 300,  "https://leetcode.com/problems/longest-increasing-subsequence/",        "DP / patience sort"),
        (17, "Edit Distance",                            "Medium", 72,   "https://leetcode.com/problems/edit-distance/",                         "2D DP"),
        (18, "Reconstruct Itinerary",                    "Hard",   332,  "https://leetcode.com/problems/reconstruct-itinerary/",                 "Hierholzer's Eulerian path"),
        (19, "Parallel Courses III",                     "Hard",   2050, "https://leetcode.com/problems/parallel-courses-iii/",                  "Topo DP — parallel execution"),
        (20, "LRU Cache",                                "Medium", 146,  "https://leetcode.com/problems/lru-cache/",                             "Cache design"),
        (21, "LFU Cache",                                "Hard",   460,  "https://leetcode.com/problems/lfu-cache/",                             "Frequency eviction"),
        (22, "Snapshot Array",                           "Medium", 1146, "https://leetcode.com/problems/snapshot-array/",                        "Binary search on history"),
        (23, "Maximum Profit Job Scheduling",            "Hard",   1235, "https://leetcode.com/problems/maximum-profit-in-job-scheduling/",      "DP + binary search"),
        (24, "Minimum Interval per Query",               "Hard",   1851, "https://leetcode.com/problems/minimum-interval-to-include-each-query/","Sort + min-heap"),
        (25, "Network Delay Time",                       "Medium", 743,  "https://leetcode.com/problems/network-delay-time/",                    "Dijkstra SSSP"),
    ],
}

# ─── Jane Street ──────────────────────────────────────────────────────────────
COMPANIES["Jane Street"] = {
    "color": "1B2A4A",
    "tips": [
        "Jane Street expects elite algorithmic problem-solving — very hard problems are common.",
        "Math and probability reasoning is tested alongside pure DSA; think like a quant.",
        "Expect OCaml or functional programming preference; but Python/C++ accepted.",
        "Recursive thinking, memoisation, and combinatorics are heavily tested.",
        "Market-making and trading strategy discussions may appear in non-coding rounds.",
    ],
    "hf": [
        (1,  "Regular Expression Matching",              "Hard",   10,   "https://leetcode.com/problems/regular-expression-matching/",           "2D DP; '*' = 0 or more"),
        (2,  "Median of Two Sorted Arrays",              "Hard",   4,    "https://leetcode.com/problems/median-of-two-sorted-arrays/",           "Binary search on partition"),
        (3,  "Trapping Rain Water",                      "Hard",   42,   "https://leetcode.com/problems/trapping-rain-water/",                   "Two pointers"),
        (4,  "Wildcard Matching",                        "Hard",   44,   "https://leetcode.com/problems/wildcard-matching/",                     "2D DP; '*' any sequence"),
        (5,  "Maximum Rectangle",                        "Hard",   85,   "https://leetcode.com/problems/maximal-rectangle/",                     "Histogram DP per row"),
        (6,  "Count of Smaller Numbers After Self",      "Hard",   315,  "https://leetcode.com/problems/count-of-smaller-numbers-after-self/",   "Merge sort / BIT / Segment tree"),
        (7,  "Largest Rectangle in Histogram",           "Hard",   84,   "https://leetcode.com/problems/largest-rectangle-in-histogram/",        "Monotonic stack"),
        (8,  "Sliding Window Maximum",                   "Hard",   239,  "https://leetcode.com/problems/sliding-window-maximum/",                "Monotonic deque"),
        (9,  "Find Median from Data Stream",             "Hard",   295,  "https://leetcode.com/problems/find-median-from-data-stream/",          "Two heaps"),
        (10, "Count of Range Sum",                       "Hard",   327,  "https://leetcode.com/problems/count-of-range-sum/",                    "Merge sort / BIT"),
        (11, "Russian Doll Envelopes",                   "Hard",   354,  "https://leetcode.com/problems/russian-doll-envelopes/",                "Sort + LIS (patience sort)"),
        (12, "Longest Increasing Subsequence",           "Medium", 300,  "https://leetcode.com/problems/longest-increasing-subsequence/",        "Patience sort O(n log n)"),
        (13, "Burst Balloons",                           "Hard",   312,  "https://leetcode.com/problems/burst-balloons/",                        "Interval DP; last balloon"),
        (14, "Super Egg Drop",                           "Hard",   887,  "https://leetcode.com/problems/super-egg-drop/",                        "DP + binary search or math"),
        (15, "Frog Jump",                                "Hard",   403,  "https://leetcode.com/problems/frog-jump/",                             "DP with set of allowed jumps"),
        (16, "Basic Calculator",                         "Hard",   224,  "https://leetcode.com/problems/basic-calculator/",                      "Stack for sign context"),
        (17, "N-Queens",                                 "Hard",   51,   "https://leetcode.com/problems/n-queens/",                              "Backtracking; col/diag sets"),
        (18, "N-Queens II",                              "Hard",   52,   "https://leetcode.com/problems/n-queens-ii/",                           "Count solutions — combinatorics"),
        (19, "Sudoku Solver",                            "Hard",   37,   "https://leetcode.com/problems/sudoku-solver/",                         "Backtracking; constraint propagation"),
        (20, "Race Car",                                 "Hard",   818,  "https://leetcode.com/problems/race-car/",                              "BFS or DP on position+speed"),
        (21, "Maximum Profit Job Scheduling",            "Hard",   1235, "https://leetcode.com/problems/maximum-profit-in-job-scheduling/",      "Sort + DP + binary search"),
        (22, "Stone Game VII",                           "Medium", 1690, "https://leetcode.com/problems/stone-game-vii/",                        "Interval DP; game theory"),
        (23, "Edit Distance",                            "Medium", 72,   "https://leetcode.com/problems/edit-distance/",                         "Classic 2D DP"),
        (24, "Reverse Pairs",                            "Hard",   493,  "https://leetcode.com/problems/reverse-pairs/",                         "Merge sort count"),
        (25, "Number of Atoms",                          "Hard",   726,  "https://leetcode.com/problems/number-of-atoms/",                       "Stack-based parsing — quant flavour"),
    ],
}

# ─── LinkedIn ─────────────────────────────────────────────────────────────────
COMPANIES["LinkedIn"] = {
    "color": "0A66C2",
    "tips": [
        "LinkedIn's graph problems model social networks — master BFS, DFS, and Union-Find.",
        "Expect nested list, iterator design, and graph connectivity questions.",
        "Job scheduling and time-series DP problems appear with financial/career context.",
        "Coding followed by design: design a newsfeed, job recommendation engine, or connection graph.",
        "Python or Java preferred; clean OOP design is valued.",
    ],
    "hf": [
        (1,  "Accounts Merge",                           "Medium", 721,  "https://leetcode.com/problems/accounts-merge/",                        "Union-Find — social graph"),
        (2,  "Maximum Profit Job Scheduling",            "Hard",   1235, "https://leetcode.com/problems/maximum-profit-in-job-scheduling/",      "Job scheduling DP"),
        (3,  "Find Median from Data Stream",             "Hard",   295,  "https://leetcode.com/problems/find-median-from-data-stream/",          "Two heaps"),
        (4,  "Nested List Weight Sum",                   "Medium", 339,  "https://leetcode.com/problems/nested-list-weight-sum/",                "DFS — LinkedIn classic"),
        (5,  "Nested List Weight Sum II",                "Medium", 364,  "https://leetcode.com/problems/nested-list-weight-sum-ii/",             "Reverse DFS weights"),
        (6,  "Serialize Deserialize Binary Tree",        "Hard",   297,  "https://leetcode.com/problems/serialize-and-deserialize-binary-tree/", "Tree codec"),
        (7,  "Max Points on a Line",                     "Hard",   149,  "https://leetcode.com/problems/max-points-on-a-line/",                  "Slope hash map"),
        (8,  "Paint House II",                           "Hard",   265,  "https://leetcode.com/problems/paint-house-ii/",                        "DP with k colours"),
        (9,  "Merge K Sorted Lists",                     "Hard",   23,   "https://leetcode.com/problems/merge-k-sorted-lists/",                  "Min-heap — feed aggregation"),
        (10, "Word Break II",                            "Hard",   140,  "https://leetcode.com/problems/word-break-ii/",                         "Backtracking + memo"),
        (11, "Number of Islands",                        "Medium", 200,  "https://leetcode.com/problems/number-of-islands/",                     "BFS/DFS — network clusters"),
        (12, "Graph Valid Tree",                         "Medium", 261,  "https://leetcode.com/problems/graph-valid-tree/",                      "Union-Find or DFS"),
        (13, "Number of Connected Components",           "Medium", 323,  "https://leetcode.com/problems/number-of-connected-components-in-an-undirected-graph/", "Union-Find"),
        (14, "Walls and Gates",                          "Medium", 286,  "https://leetcode.com/problems/walls-and-gates/",                       "Multi-source BFS"),
        (15, "Minimum Spanning Tree — Connect Points",  "Medium", 1584, "https://leetcode.com/problems/min-cost-to-connect-all-points/",        "Prim's or Kruskal's"),
        (16, "Flatten Nested List Iterator",             "Medium", 341,  "https://leetcode.com/problems/flatten-nested-list-iterator/",          "Stack-based iterator"),
        (17, "Random Pick with Weight",                  "Medium", 528,  "https://leetcode.com/problems/random-pick-with-weight/",               "Prefix sum + binary search"),
        (18, "Maximum Swap",                             "Medium", 670,  "https://leetcode.com/problems/maximum-swap/",                          "Greedy last occurrence"),
        (19, "Two Sum III — Data Structure Design",      "Easy",   170,  "https://leetcode.com/problems/two-sum-iii-data-structure-design/",     "Hash map design"),
        (20, "Sparse Matrix Multiplication",             "Medium", 311,  "https://leetcode.com/problems/sparse-matrix-multiplication/",          "Sparse optimisation"),
        (21, "Closest Binary Search Tree Value",         "Easy",   270,  "https://leetcode.com/problems/closest-binary-search-tree-value/",      "BST traversal"),
        (22, "Alien Dictionary",                         "Hard",   269,  "https://leetcode.com/problems/alien-dictionary/",                      "Topo sort from char order"),
        (23, "Find Leaves of Binary Tree",               "Medium", 366,  "https://leetcode.com/problems/find-leaves-of-binary-tree/",            "Height-based grouping"),
        (24, "Maximum Product of Word Lengths",          "Medium", 318,  "https://leetcode.com/problems/maximum-product-of-word-lengths/",       "Bit mask for char sets"),
        (25, "Minimum Height Trees",                     "Medium", 310,  "https://leetcode.com/problems/minimum-height-trees/",                  "Topological leaf trimming"),
    ],
}

# ─── Snowflake ────────────────────────────────────────────────────────────────
COMPANIES["Snowflake"] = {
    "color": "29B5E8",
    "tips": [
        "Snowflake interviews test data structure fundamentals with cloud/data-warehouse context.",
        "Expect system design around column-store databases, caching, and query processing.",
        "Coding rounds focus on arrays, heaps, graphs, and design questions.",
        "SQL query optimisation knowledge is a plus for data engineering roles.",
        "Demonstrate scalability thinking: how does your solution perform at petabyte scale?",
    ],
    "hf": [
        (1,  "Two Sum",                                  "Easy",   1,    "https://leetcode.com/problems/two-sum/",                               "Hash map"),
        (2,  "Longest Substring Without Repeating",      "Medium", 3,    "https://leetcode.com/problems/longest-substring-without-repeating-characters/", "Sliding window"),
        (3,  "Median of Two Sorted Arrays",              "Hard",   4,    "https://leetcode.com/problems/median-of-two-sorted-arrays/",           "Binary search — analytics"),
        (4,  "Valid Parentheses",                        "Easy",   20,   "https://leetcode.com/problems/valid-parentheses/",                     "Stack"),
        (5,  "Merge Two Sorted Lists",                   "Easy",   21,   "https://leetcode.com/problems/merge-two-sorted-lists/",                "Dummy head"),
        (6,  "Maximum Subarray",                         "Medium", 53,   "https://leetcode.com/problems/maximum-subarray/",                      "Kadane's"),
        (7,  "Number of Islands",                        "Medium", 200,  "https://leetcode.com/problems/number-of-islands/",                     "BFS/DFS"),
        (8,  "Course Schedule",                          "Medium", 207,  "https://leetcode.com/problems/course-schedule/",                       "Topo sort"),
        (9,  "Clone Graph",                              "Medium", 133,  "https://leetcode.com/problems/clone-graph/",                           "BFS + hash map"),
        (10, "Word Break",                               "Medium", 139,  "https://leetcode.com/problems/word-break/",                            "1D DP"),
        (11, "Merge Intervals",                          "Medium", 56,   "https://leetcode.com/problems/merge-intervals/",                       "Sort + merge"),
        (12, "LRU Cache",                                "Medium", 146,  "https://leetcode.com/problems/lru-cache/",                             "Cache design — warehouse context"),
        (13, "Evaluate Reverse Polish Notation",         "Medium", 150,  "https://leetcode.com/problems/evaluate-reverse-polish-notation/",      "Stack — query evaluation"),
        (14, "Coin Change",                              "Medium", 322,  "https://leetcode.com/problems/coin-change/",                           "Unbounded knapsack DP"),
        (15, "Top K Frequent Elements",                  "Medium", 347,  "https://leetcode.com/problems/top-k-frequent-elements/",               "Heap"),
        (16, "Accounts Merge",                           "Medium", 721,  "https://leetcode.com/problems/accounts-merge/",                        "Union-Find — data dedup"),
        (17, "Find Median from Data Stream",             "Hard",   295,  "https://leetcode.com/problems/find-median-from-data-stream/",          "Two heaps — streaming analytics"),
        (18, "Sliding Window Maximum",                   "Hard",   239,  "https://leetcode.com/problems/sliding-window-maximum/",                "Monotonic deque"),
        (19, "Word Ladder",                              "Hard",   127,  "https://leetcode.com/problems/word-ladder/",                           "BFS"),
        (20, "Critical Connections in Network",          "Hard",   1192, "https://leetcode.com/problems/critical-connections-in-a-network/",     "Tarjan's bridge"),
        (21, "Design Underground System",                "Medium", 1396, "https://leetcode.com/problems/design-underground-system/",             "Two hash maps — query tracking"),
        (22, "Snapshot Array",                           "Medium", 1146, "https://leetcode.com/problems/snapshot-array/",                        "Binary search on history — versioning"),
        (23, "Maximum Profit Job Scheduling",            "Hard",   1235, "https://leetcode.com/problems/maximum-profit-in-job-scheduling/",      "DP + binary search"),
        (24, "Minimum Interval per Query",               "Hard",   1851, "https://leetcode.com/problems/minimum-interval-to-include-each-query/","Sort + min-heap"),
        (25, "Alien Dictionary",                         "Hard",   269,  "https://leetcode.com/problems/alien-dictionary/",                      "Topo sort — schema ordering"),
    ],
}

# ─── Meta (Facebook) ─────────────────────────────────────────────────────────
COMPANIES["Meta"] = {
    "color": "0668E1",
    "tips": [
        "Meta values fast, iterative coding. Expect 2 coding rounds + system design + behavioural.",
        "Top topics: Arrays, Strings, Trees, Graphs. Meta has unique 'signal' problems.",
        "System design tests social graph, news feed ranking, Messenger, and Instagram-scale systems.",
        "Behavioural: Meta core values — Move Fast, Be Bold, Focus on Impact, Be Open, Build Social Value.",
        "Master problems with parentheses manipulation, interval merging, and graph connectivity.",
    ],
    "hf": [
        (1,  "Merge Intervals",                          "Medium", 56,   "https://leetcode.com/problems/merge-intervals/",                       "Meta top-asked"),
        (2,  "Valid Parentheses",                        "Easy",   20,   "https://leetcode.com/problems/valid-parentheses/",                     "Stack"),
        (3,  "Binary Tree Right Side View",              "Medium", 199,  "https://leetcode.com/problems/binary-tree-right-side-view/",           "BFS last per level"),
        (4,  "Minimum Remove to Make Valid Parens",      "Medium", 1249, "https://leetcode.com/problems/minimum-remove-to-make-valid-parentheses/","Stack of indices"),
        (5,  "Move Zeroes",                              "Easy",   283,  "https://leetcode.com/problems/move-zeroes/",                           "In-place two pointers"),
        (6,  "Product of Array Except Self",             "Medium", 238,  "https://leetcode.com/problems/product-of-array-except-self/",          "Prefix & suffix"),
        (7,  "Subarray Sum Equals K",                    "Medium", 560,  "https://leetcode.com/problems/subarray-sum-equals-k/",                 "Prefix sum + hash map"),
        (8,  "Add Binary",                               "Easy",   67,   "https://leetcode.com/problems/add-binary/",                            "Bit carry simulation"),
        (9,  "LCA of Binary Tree",                       "Medium", 236,  "https://leetcode.com/problems/lowest-common-ancestor-of-a-binary-tree/","Post-order DFS"),
        (10, "Accounts Merge",                           "Medium", 721,  "https://leetcode.com/problems/accounts-merge/",                        "Union-Find — social graph"),
        (11, "Number of Islands",                        "Medium", 200,  "https://leetcode.com/problems/number-of-islands/",                     "BFS/DFS"),
        (12, "Merge K Sorted Lists",                     "Hard",   23,   "https://leetcode.com/problems/merge-k-sorted-lists/",                  "Min-heap — feed merge"),
        (13, "Remove Invalid Parentheses",               "Hard",   301,  "https://leetcode.com/problems/remove-invalid-parentheses/",            "BFS all min removals"),
        (14, "Flatten Binary Tree to Linked List",       "Medium", 114,  "https://leetcode.com/problems/flatten-binary-tree-to-linked-list/",    "Morris traversal"),
        (15, "3Sum",                                     "Medium", 15,   "https://leetcode.com/problems/3sum/",                                  "Sort + two pointers"),
        (16, "Task Scheduler",                           "Medium", 621,  "https://leetcode.com/problems/task-scheduler/",                        "Greedy + cooldown"),
        (17, "Custom Sort String",                       "Medium", 791,  "https://leetcode.com/problems/custom-sort-string/",                    "Sort with custom key"),
        (18, "Interval List Intersections",              "Medium", 986,  "https://leetcode.com/problems/interval-list-intersections/",           "Two pointers on both lists"),
        (19, "Binary Tree Vertical Order Traversal",     "Medium", 314,  "https://leetcode.com/problems/binary-tree-vertical-order-traversal/",  "BFS + col tracking"),
        (20, "Buildings With an Ocean View",             "Medium", 1762, "https://leetcode.com/problems/buildings-with-an-ocean-view/",          "Monotonic stack from right"),
        (21, "Minimum Window Substring",                 "Hard",   76,   "https://leetcode.com/problems/minimum-window-substring/",              "Sliding window"),
        (22, "Word Search II",                           "Hard",   212,  "https://leetcode.com/problems/word-search-ii/",                        "Trie + DFS"),
        (23, "Median from Data Stream",                  "Hard",   295,  "https://leetcode.com/problems/find-median-from-data-stream/",          "Two heaps"),
        (24, "Fraction to Recurring Decimal",            "Medium", 166,  "https://leetcode.com/problems/fraction-to-recurring-decimal/",         "Hash map for cycle detect"),
        (25, "Next Permutation",                         "Medium", 31,   "https://leetcode.com/problems/next-permutation/",                      "Find pivot + reverse suffix"),
    ],
}

# ─── Apple ───────────────────────────────────────────────────────────────────
COMPANIES["Apple"] = {
    "color": "555555",
    "tips": [
        "Apple interviews are thorough — expect 5–6 rounds covering coding, design, and culture fit.",
        "Attention to edge cases and correctness is paramount at Apple.",
        "Design questions often focus on iOS-scale APIs, offline-first systems, and media pipelines.",
        "Trees, DP, arrays, and design problems are the top areas.",
        "Demonstrate deep understanding of algorithms — Apple values quality over speed.",
    ],
    "hf": [
        (1,  "Two Sum",                                  "Easy",   1,    "https://leetcode.com/problems/two-sum/",                               "Hash map"),
        (2,  "Merge Two Sorted Lists",                   "Easy",   21,   "https://leetcode.com/problems/merge-two-sorted-lists/",                "Dummy head"),
        (3,  "Maximum Subarray",                         "Medium", 53,   "https://leetcode.com/problems/maximum-subarray/",                      "Kadane's"),
        (4,  "LRU Cache",                                "Medium", 146,  "https://leetcode.com/problems/lru-cache/",                             "Design — cache eviction"),
        (5,  "Number of Islands",                        "Medium", 200,  "https://leetcode.com/problems/number-of-islands/",                     "BFS/DFS"),
        (6,  "Flatten Binary Tree to Linked List",       "Medium", 114,  "https://leetcode.com/problems/flatten-binary-tree-to-linked-list/",    "Morris traversal"),
        (7,  "Longest Increasing Subsequence",           "Medium", 300,  "https://leetcode.com/problems/longest-increasing-subsequence/",        "DP or patience sort"),
        (8,  "Word Break",                               "Medium", 139,  "https://leetcode.com/problems/word-break/",                            "1D DP"),
        (9,  "3Sum",                                     "Medium", 15,   "https://leetcode.com/problems/3sum/",                                  "Sort + two pointers"),
        (10, "Longest Substring Without Repeating",      "Medium", 3,    "https://leetcode.com/problems/longest-substring-without-repeating-characters/","Sliding window"),
        (11, "Find Minimum in Rotated Sorted Array",     "Medium", 153,  "https://leetcode.com/problems/find-minimum-in-rotated-sorted-array/",  "Binary search"),
        (12, "Course Schedule",                          "Medium", 207,  "https://leetcode.com/problems/course-schedule/",                       "Topo sort"),
        (13, "Binary Tree Right Side View",              "Medium", 199,  "https://leetcode.com/problems/binary-tree-right-side-view/",           "BFS last per level"),
        (14, "Validate Binary Search Tree",              "Medium", 98,   "https://leetcode.com/problems/validate-binary-search-tree/",           "Min/max range DFS"),
        (15, "Count of Range Sum",                       "Hard",   327,  "https://leetcode.com/problems/count-of-range-sum/",                    "Merge sort / BIT"),
        (16, "Maximum Rectangle",                        "Hard",   85,   "https://leetcode.com/problems/maximal-rectangle/",                     "Histogram DP"),
        (17, "Median of Two Sorted Arrays",              "Hard",   4,    "https://leetcode.com/problems/median-of-two-sorted-arrays/",           "Binary search on partition"),
        (18, "Trapping Rain Water",                      "Hard",   42,   "https://leetcode.com/problems/trapping-rain-water/",                   "Two pointers"),
        (19, "Serialize Deserialize Binary Tree",        "Hard",   297,  "https://leetcode.com/problems/serialize-and-deserialize-binary-tree/", "BFS or preorder"),
        (20, "Design In-Memory File System",             "Hard",   588,  "https://leetcode.com/problems/design-in-memory-file-system/",          "Trie OOP — Apple iOS filesystem"),
        (21, "Edit Distance",                            "Medium", 72,   "https://leetcode.com/problems/edit-distance/",                         "2D DP"),
        (22, "Binary Tree Maximum Path Sum",             "Hard",   124,  "https://leetcode.com/problems/binary-tree-maximum-path-sum/",          "Post-order; global max"),
        (23, "Word Search II",                           "Hard",   212,  "https://leetcode.com/problems/word-search-ii/",                        "Trie + DFS"),
        (24, "Accounts Merge",                           "Medium", 721,  "https://leetcode.com/problems/accounts-merge/",                        "Union-Find"),
        (25, "My Calendar I",                            "Medium", 729,  "https://leetcode.com/problems/my-calendar-i/",                         "Binary search on sorted intervals"),
    ],
}

# ─── HRT (Hudson River Trading) ───────────────────────────────────────────────
COMPANIES["HRT"] = {
    "color": "C8102E",
    "tips": [
        "HRT expects elite competitive-programming-level problem solving — Hard problems are standard.",
        "Mathematical rigour is key: probability, combinatorics, and number theory appear frequently.",
        "Brain teasers and market-making puzzles supplement coding rounds.",
        "Know advanced data structures: segment trees, BITs, persistent data structures.",
        "C++ performance expertise is highly valued; algorithmic trading latency matters.",
    ],
    "hf": [
        (1,  "Regular Expression Matching",              "Hard",   10,   "https://leetcode.com/problems/regular-expression-matching/",           "2D DP"),
        (2,  "Median of Two Sorted Arrays",              "Hard",   4,    "https://leetcode.com/problems/median-of-two-sorted-arrays/",           "Binary search on partition"),
        (3,  "Trapping Rain Water",                      "Hard",   42,   "https://leetcode.com/problems/trapping-rain-water/",                   "Two pointers"),
        (4,  "Maximum Rectangle",                        "Hard",   85,   "https://leetcode.com/problems/maximal-rectangle/",                     "Histogram DP"),
        (5,  "Largest Rectangle in Histogram",           "Hard",   84,   "https://leetcode.com/problems/largest-rectangle-in-histogram/",        "Monotonic stack"),
        (6,  "Sliding Window Maximum",                   "Hard",   239,  "https://leetcode.com/problems/sliding-window-maximum/",                "Monotonic deque"),
        (7,  "Count of Smaller Numbers After Self",      "Hard",   315,  "https://leetcode.com/problems/count-of-smaller-numbers-after-self/",   "Merge sort / BIT / Segment tree"),
        (8,  "Count of Range Sum",                       "Hard",   327,  "https://leetcode.com/problems/count-of-range-sum/",                    "Merge sort / BIT — quant style"),
        (9,  "Russian Doll Envelopes",                   "Hard",   354,  "https://leetcode.com/problems/russian-doll-envelopes/",                "Sort + LIS patience sort"),
        (10, "Burst Balloons",                           "Hard",   312,  "https://leetcode.com/problems/burst-balloons/",                        "Interval DP"),
        (11, "Super Egg Drop",                           "Hard",   887,  "https://leetcode.com/problems/super-egg-drop/",                        "DP + binary search or math"),
        (12, "Race Car",                                 "Hard",   818,  "https://leetcode.com/problems/race-car/",                              "BFS or DP on state"),
        (13, "Find Median from Data Stream",             "Hard",   295,  "https://leetcode.com/problems/find-median-from-data-stream/",          "Two heaps — market mid-price"),
        (14, "Minimum Window Substring",                 "Hard",   76,   "https://leetcode.com/problems/minimum-window-substring/",              "Sliding window"),
        (15, "Word Ladder",                              "Hard",   127,  "https://leetcode.com/problems/word-ladder/",                           "BFS"),
        (16, "N-Queens",                                 "Hard",   51,   "https://leetcode.com/problems/n-queens/",                              "Backtracking"),
        (17, "Sudoku Solver",                            "Hard",   37,   "https://leetcode.com/problems/sudoku-solver/",                         "Backtracking + constraint propagation"),
        (18, "Longest Increasing Subsequence",           "Medium", 300,  "https://leetcode.com/problems/longest-increasing-subsequence/",        "Patience sort O(n log n)"),
        (19, "Edit Distance",                            "Medium", 72,   "https://leetcode.com/problems/edit-distance/",                         "2D DP"),
        (20, "Stone Game VII",                           "Medium", 1690, "https://leetcode.com/problems/stone-game-vii/",                        "Interval DP — game theory"),
        (21, "Reverse Pairs",                            "Hard",   493,  "https://leetcode.com/problems/reverse-pairs/",                         "Merge sort count — quant context"),
        (22, "Maximum Profit Job Scheduling",            "Hard",   1235, "https://leetcode.com/problems/maximum-profit-in-job-scheduling/",      "DP + binary search"),
        (23, "Range Sum Query — Mutable",                "Medium", 307,  "https://leetcode.com/problems/range-sum-query-mutable/",               "Segment tree or BIT"),
        (24, "Count Different Palindromic Subsequences", "Hard",   730,  "https://leetcode.com/problems/count-different-palindromic-subsequences/","DP — counting"),
        (25, "Minimum Number of Refueling Stops",        "Hard",   871,  "https://leetcode.com/problems/minimum-number-of-refueling-stops/",     "Greedy + max-heap"),
    ],
}

# ─── OpenAI ───────────────────────────────────────────────────────────────────
COMPANIES["OpenAI"] = {
    "color": "10A37F",
    "tips": [
        "OpenAI interviews test strong fundamentals plus ML/AI-adjacent algorithmic thinking.",
        "Expect graph, DP, and design problems at Medium–Hard difficulty.",
        "System design covers ML pipelines, model serving, distributed training infra.",
        "Curiosity and research mindset are valued — discuss alternative approaches.",
        "Python proficiency is essential; familiarity with PyTorch/NumPy internals is a plus.",
    ],
    "hf": [
        (1,  "Two Sum",                                  "Easy",   1,    "https://leetcode.com/problems/two-sum/",                               "Hash map"),
        (2,  "Valid Parentheses",                        "Easy",   20,   "https://leetcode.com/problems/valid-parentheses/",                     "Stack"),
        (3,  "Number of Islands",                        "Medium", 200,  "https://leetcode.com/problems/number-of-islands/",                     "BFS/DFS"),
        (4,  "Clone Graph",                              "Medium", 133,  "https://leetcode.com/problems/clone-graph/",                           "BFS + hash map"),
        (5,  "Course Schedule",                          "Medium", 207,  "https://leetcode.com/problems/course-schedule/",                       "Topo sort — model dependency"),
        (6,  "Word Break",                               "Medium", 139,  "https://leetcode.com/problems/word-break/",                            "1D DP — tokenisation"),
        (7,  "Longest Increasing Subsequence",           "Medium", 300,  "https://leetcode.com/problems/longest-increasing-subsequence/",        "DP / patience sort"),
        (8,  "Edit Distance",                            "Medium", 72,   "https://leetcode.com/problems/edit-distance/",                         "2D DP — text similarity"),
        (9,  "Top K Frequent Elements",                  "Medium", 347,  "https://leetcode.com/problems/top-k-frequent-elements/",               "Heap"),
        (10, "Implement Trie",                           "Medium", 208,  "https://leetcode.com/problems/implement-trie-prefix-tree/",            "Trie — autocomplete"),
        (11, "Find Median from Data Stream",             "Hard",   295,  "https://leetcode.com/problems/find-median-from-data-stream/",          "Two heaps — streaming"),
        (12, "Word Search II",                           "Hard",   212,  "https://leetcode.com/problems/word-search-ii/",                        "Trie + DFS"),
        (13, "Serialize Deserialize Binary Tree",        "Hard",   297,  "https://leetcode.com/problems/serialize-and-deserialize-binary-tree/", "Model checkpoint serialisation"),
        (14, "Alien Dictionary",                         "Hard",   269,  "https://leetcode.com/problems/alien-dictionary/",                      "Topo sort — language modelling"),
        (15, "Trapping Rain Water",                      "Hard",   42,   "https://leetcode.com/problems/trapping-rain-water/",                   "Two pointers"),
        (16, "Median of Two Sorted Arrays",              "Hard",   4,    "https://leetcode.com/problems/median-of-two-sorted-arrays/",           "Binary search on partition"),
        (17, "Decode Ways",                              "Medium", 91,   "https://leetcode.com/problems/decode-ways/",                           "1D DP — tokenisation context"),
        (18, "Regular Expression Matching",              "Hard",   10,   "https://leetcode.com/problems/regular-expression-matching/",           "2D DP — pattern matching"),
        (19, "N-Queens",                                 "Hard",   51,   "https://leetcode.com/problems/n-queens/",                              "Backtracking — constraint solving"),
        (20, "Burst Balloons",                           "Hard",   312,  "https://leetcode.com/problems/burst-balloons/",                        "Interval DP"),
        (21, "Accounts Merge",                           "Medium", 721,  "https://leetcode.com/problems/accounts-merge/",                        "Union-Find"),
        (22, "Coin Change",                              "Medium", 322,  "https://leetcode.com/problems/coin-change/",                           "DP — resource allocation"),
        (23, "Binary Tree Maximum Path Sum",             "Hard",   124,  "https://leetcode.com/problems/binary-tree-maximum-path-sum/",          "Post-order; global max"),
        (24, "Task Scheduler",                           "Medium", 621,  "https://leetcode.com/problems/task-scheduler/",                        "Greedy — GPU scheduling analogy"),
        (25, "Maximum Profit Job Scheduling",            "Hard",   1235, "https://leetcode.com/problems/maximum-profit-in-job-scheduling/",      "DP + binary search"),
    ],
}


# ═══════════════════════════════════════════════════════════════════════════════
#  GENERATION
# ═══════════════════════════════════════════════════════════════════════════════

TOPIC_ORDER = [
    "Arrays & Hashing",
    "Two Pointers",
    "Sliding Window",
    "Stack",
    "Binary Search",
    "Linked List",
    "Trees",
    "Heap / Priority Queue",
    "Backtracking",
    "Tries",
    "Graphs",
    "DP — 1D",
    "DP — 2D",
    "Greedy",
    "Intervals",
    "Math & Geometry",
    "Bit Manipulation",
    "Design / OOP",
    "String Manipulation",
]

# Focus level per topic (used in overview)
TOPIC_FOCUS = {
    "Arrays & Hashing"   : "★★★★☆  Core",
    "Two Pointers"        : "★★★★☆  High",
    "Sliding Window"      : "★★★★☆  High",
    "Stack"               : "★★★☆☆  Medium",
    "Binary Search"       : "★★★★☆  High",
    "Linked List"         : "★★★☆☆  Medium",
    "Trees"               : "★★★★★  Very High",
    "Heap / Priority Queue": "★★★★☆  High",
    "Backtracking"        : "★★★★☆  High",
    "Tries"               : "★★★☆☆  Medium",
    "Graphs"              : "★★★★★  Very High",
    "DP — 1D"             : "★★★★★  Highest",
    "DP — 2D"             : "★★★★★  Highest",
    "Greedy"              : "★★★★☆  High",
    "Intervals"           : "★★★★☆  High",
    "Math & Geometry"     : "★★★☆☆  Medium",
    "Bit Manipulation"    : "★★★☆☆  Medium",
    "Design / OOP"        : "★★★★☆  High",
    "String Manipulation" : "★★★★☆  High",
}


def generate_company_workbook(company_name, company_data):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    color = company_data["color"]
    tips  = company_data["tips"]
    hf    = company_data["hf"]

    # Overview entries: (sheet_name, q_count, topics_str, priority, note)
    overview_entries = [("⭐ High Frequency", len(hf), "Must-know questions", "★★★★★  PRIORITY", "Start here!")]

    for topic in TOPIC_ORDER:
        rows = TOPICS.get(topic, [])
        if not rows:
            continue
        add_topic_sheet(wb, topic, rows, color, company_name)
        overview_entries.append((
            safe_title(topic),
            len(rows),
            topic,
            TOPIC_FOCUS.get(topic, "★★★☆☆"),
            f"{sum(1 for r in rows if r[2]=='Easy')}E / "
            f"{sum(1 for r in rows if r[2]=='Medium')}M / "
            f"{sum(1 for r in rows if r[2]=='Hard')}H",
        ))

    # High Frequency sheet (added last in WB, placed after topics)
    add_hf_sheet(wb, company_name, color, hf)

    # Move High Frequency to position 0 (after overview)
    hf_ws = wb["⭐ High Frequency"]
    wb.move_sheet(hf_ws, offset=-len(wb.sheetnames) + 1)

    # Overview sheet (index 0)
    add_overview_sheet(wb, company_name, color, tips, overview_entries)

    # Save
    safe_name = company_name.replace(" ", "_").replace("/", "-")
    out_path = os.path.join(OUT_DIR, f"{safe_name}_DSA_Interview_Prep.xlsx")
    wb.save(out_path)
    total_q = sum(len(TOPICS.get(t, [])) for t in TOPIC_ORDER) + len(hf)
    print(f"  ✅  {company_name:<20} → {os.path.basename(out_path)}  ({total_q} questions, {len(wb.sheetnames)} sheets)")
    return out_path


# ─── Main ─────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print(f"\n{'='*65}")
    print(f"  DSA Interview Prep Excel Generator — 16 Companies")
    print(f"  Output folder: {OUT_DIR}")
    print(f"{'='*65}\n")

    generated = []
    for company_name, company_data in COMPANIES.items():
        path = generate_company_workbook(company_name, company_data)
        generated.append(path)

    print(f"\n{'='*65}")
    print(f"  Generated {len(generated)} workbooks successfully!")
    print(f"  Each workbook has {len(TOPIC_ORDER)} topic sheets + 1 High Frequency sheet + 1 Overview sheet.")
    print(f"{'='*65}\n")
